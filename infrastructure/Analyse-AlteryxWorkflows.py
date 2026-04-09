"""
Alteryx Workflow Migration Analyser
===================================
Parses exported .yxzp/.yxmd files and produces an Excel migration inventory.

Usage:
    python Analyse-AlteryxWorkflows.py C:\AlteryxExport\workflows C:\AlteryxExport\migration_report.xlsx

Requirements:
    pip install openpyxl lxml
"""

import sys, os, zipfile, tempfile, re, shutil
from pathlib import Path
from collections import Counter, defaultdict
from lxml import etree
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── Tool classification ───────────────────────────────────────────────────────

PLUGIN_MAP = {
    "DbFileInput": "DB Input",
    "DbFileOutput": "DB Output",
    "AlteryxSelect": "Select",
    "Filter": "Filter",
    "Formula": "Formula",
    "MultiFieldFormula": "Multi-Field Formula",
    "MultiRowFormula": "Multi-Row Formula",
    "Sort": "Sort",
    "Unique": "Unique",
    "Sample": "Sample",
    "DataCleansing": "Data Cleansing",
    "RecordID": "Record ID",
    "GenerateRows": "Generate Rows",
    "DateTime": "DateTime",
    "RunningTotal": "Running Total",
    "Join": "Join",
    "FindReplace": "Find Replace",
    "AppendFields": "Append Fields",
    "Union": "Union",
    "JoinMultiple": "Join Multiple",
    "CrossTab": "Cross Tab",
    "MakeCrossTab": "Cross Tab",
    "Transpose": "Transpose",
    "Summarize": "Summarize",
    "RegEx": "RegEx",
    "XMLParse": "XML Parse",
    "JSONParse": "JSON Parse",
    "TextToColumns": "Text to Columns",
    "ColumnToRows": "Column to Rows",
    "SpatialMatch": "Spatial Match",
    "TradeArea": "Trade Area",
    "Distance": "Distance",
    "Geocoder": "Geocoder",
    "MapInput": "Map Input",
    "PolyBuild": "Poly Build",
    "PolySplit": "Poly Split",
    "SpatialInfo": "Spatial Info",
    "RPluginTool": "R Tool",
    "PythonTool": "Python Tool",
    "LinearRegression": "Linear Regression",
    "LogisticRegression": "Logistic Regression",
    "DecisionTree": "Decision Tree",
    "RandomForest": "Random Forest",
    "ScoreRecords": "Score Records",
    "ToolContainer": "Tool Container",
    "DynamicInput": "Dynamic Input",
    "DynamicReplace": "Dynamic Replace",
    "RunCommand": "Run Command",
    "BlockUntilDone": "Block Until Done",
    "BatchMacro": "Batch Macro",
    "IterativeMacro": "Iterative Macro",
    "MacroInput": "Macro Input",
    "MacroOutput": "Macro Output",
    "DownloadTool": "Download (HTTP)",
    "BrowseV2": "Browse",
    "FileInput": "File Input",
    "FileOutput": "File Output",
    "TextInput": "Text Input",
    "EmailTool": "Email",
    "ReportText": "Report Text",
    "ReportHeader": "Report Header",
    "Layout": "Report Layout",
    "Render": "Report Render",
    "Table": "Report Table",
    "Chart": "Report Chart",
    "Map": "Report Map",
    "Tab": "Interface Tab",
    "TextBox": "Interface TextBox",
    "DropDown": "Interface Dropdown",
    "CheckBox": "Interface Checkbox",
    "Action": "Interface Action",
    "Date": "Interface Date",
    "NumericUpDown": "Interface Numeric",
}

SPATIAL_TOOLS = {"Spatial Match", "Trade Area", "Distance", "Geocoder", "Map Input",
                 "Poly Build", "Poly Split", "Spatial Info", "Report Map"}
PREDICTIVE_TOOLS = {"R Tool", "Python Tool", "Linear Regression", "Logistic Regression",
                    "Decision Tree", "Random Forest", "Score Records"}
INTERFACE_TOOLS = {"Interface Tab", "Interface TextBox", "Interface Dropdown",
                   "Interface Checkbox", "Interface Action", "Interface Date", "Interface Numeric"}
ORCHESTRATION_TOOLS = {"Run Command", "Block Until Done", "Batch Macro", "Iterative Macro",
                       "Dynamic Input", "Dynamic Replace", "Download (HTTP)"}


def classify_plugin(plugin_str):
    if not plugin_str:
        return "Unknown"
    for key, name in PLUGIN_MAP.items():
        if key in plugin_str:
            return name
    short = plugin_str.rsplit(".", 1)[-1] if "." in plugin_str else plugin_str
    return short


def classify_tier(tools_used, has_macros, is_macro=False):
    tool_set = set(tools_used)
    if tool_set & SPATIAL_TOOLS or tool_set & PREDICTIVE_TOOLS:
        return "Tier 3 - Predictive/Spatial/Code"
    # Macros inherently have interface elements — don't count those as self-service
    if not is_macro and tool_set & INTERFACE_TOOLS:
        return "Tier 4 - Self-Service/Interface"
    if tool_set & ORCHESTRATION_TOOLS or has_macros:
        return "Tier 2 - Transform & Orchestrate"
    return "Tier 1 - Simple ETL"


def parse_workflow(xml_path, is_macro=False):
    """Parse a single .yxmd/.yxmc file and return a structured dict."""
    result = {
        "file": str(xml_path),
        "name": "",
        "version": "",
        "is_macro": is_macro,
        "tool_count": 0,
        "tools": Counter(),
        "connections": [],
        "sql_queries": [],
        "macros": [],
        "containers": [],
        "credentials": [],
        "tier": "",
        "errors": [],
    }

    try:
        tree = etree.parse(str(xml_path))
        root = tree.getroot()
    except Exception as e:
        result["errors"].append(f"Parse error: {e}")
        return result

    # Workflow version and name
    result["version"] = root.get("yxmdVer", "")
    meta = root.find(".//MetaInfo")
    if meta is not None:
        name_el = meta.find("Name")
        if name_el is not None and name_el.text:
            result["name"] = name_el.text.strip()

    if not result["name"]:
        result["name"] = Path(xml_path).stem

    # Parse all nodes
    nodes = root.findall(".//Node")
    result["tool_count"] = len(nodes)

    for node in nodes:
        gui = node.find(".//GuiSettings")
        if gui is not None:
            plugin = gui.get("Plugin", "")
            tool_name = classify_plugin(plugin)
            result["tools"][tool_name] += 1

        # Tool containers (labelled groups)
        config = node.find(".//Configuration")
        if config is not None:
            caption = config.find("Caption")
            if caption is not None and caption.text:
                result["containers"].append(caption.text.strip())

        # Data connections
        for val in node.findall(".//Value"):
            val_name = val.get("name", "")
            val_text = val.text or ""
            if "Existing Connection" in val_name and val_text:
                result["connections"].append(val_text)
            if "Credentials" in val_name and "username" in val_name and val_text:
                result["credentials"].append(val_text)

        # Cached connection names
        for cached in node.findall(".//CachedCosmeticName"):
            if cached.text:
                result["connections"].append(f"CachedConn: {cached.text.strip()}")

        # SQL queries from <File> elements
        for file_el in node.findall(".//File"):
            if file_el.text and ("Select" in file_el.text or "FROM" in file_el.text.upper()):
                sql = file_el.text.strip()
                if len(sql) > 500:
                    sql = sql[:500] + "..."
                result["sql_queries"].append(sql)

        # Macros
        engine = node.find(".//EngineSettings")
        if engine is not None:
            macro = engine.get("Macro", "")
            if macro:
                result["macros"].append(macro)

    # Deduplicate connections
    result["connections"] = list(set(result["connections"]))
    result["credentials"] = list(set(result["credentials"]))

    # Classify tier
    result["tier"] = classify_tier(list(result["tools"].keys()), len(result["macros"]) > 0, is_macro)

    return result


def find_yxmd_files(source_dir):
    """Find all workflow files, extracting from .yxzp archives if needed."""
    source = Path(source_dir)
    workflow_exts = ["*.yxmd", "*.yxwz", "*.yxwg"]
    macro_exts = ["*.yxmc"]

    workflow_files = []
    for ext in workflow_exts:
        workflow_files.extend(source.rglob(ext))

    yxmc_files = []
    for ext in macro_exts:
        yxmc_files.extend(source.rglob(ext))

    temp_dirs = []
    for yxzp in source.rglob("*.yxzp"):
        tmp = tempfile.mkdtemp(prefix="alteryx_")
        temp_dirs.append(tmp)
        try:
            with zipfile.ZipFile(str(yxzp), "r") as zf:
                zf.extractall(tmp)
            for ext in workflow_exts:
                workflow_files.extend(Path(tmp).rglob(ext))
            for ext in macro_exts:
                yxmc_files.extend(Path(tmp).rglob(ext))
        except Exception as e:
            print(f"  WARN: Could not extract {yxzp.name}: {e}")

    return workflow_files, yxmc_files, temp_dirs


def build_report(results, output_path):
    """Generate the Excel migration report."""
    wb = Workbook()

    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill("solid", fgColor="2F5496")
    tier1_fill = PatternFill("solid", fgColor="C6EFCE")
    tier2_fill = PatternFill("solid", fgColor="FFEB9C")
    tier3_fill = PatternFill("solid", fgColor="FFC7CE")
    tier4_fill = PatternFill("solid", fgColor="D9E1F2")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )

    tier_fills = {
        "Tier 1 - Simple ETL": tier1_fill,
        "Tier 2 - Transform & Orchestrate": tier2_fill,
        "Tier 3 - Predictive/Spatial/Code": tier3_fill,
        "Tier 4 - Self-Service/Interface": tier4_fill,
    }

    # ── Sheet 1: Executive Summary ────────────────────────────────────────
    ws = wb.active
    ws.title = "Summary"

    tier_counts = Counter(r["tier"] for r in results if r["tier"] and not r.get("is_macro"))
    macro_tier_counts = Counter(r["tier"] for r in results if r["tier"] and r.get("is_macro"))
    wf_total = sum(1 for r in results if not r.get("is_macro"))
    macro_total = sum(1 for r in results if r.get("is_macro"))
    total = len(results)

    summary_data = [
        ["Alteryx Migration Inventory - Executive Summary", ""],
        ["", ""],
        ["Total Items Analysed", total],
        ["  Workflows", wf_total],
        ["  Macros", macro_total],
        ["", ""],
        ["Workflow Tier Breakdown", "Count"],
    ]
    for tier in ["Tier 1 - Simple ETL", "Tier 2 - Transform & Orchestrate",
                 "Tier 3 - Predictive/Spatial/Code", "Tier 4 - Self-Service/Interface"]:
        summary_data.append([tier, tier_counts.get(tier, 0)])

    summary_data += [
        ["", ""],
        ["Macro Tier Breakdown", "Count"],
    ]
    for tier in ["Tier 1 - Simple ETL", "Tier 2 - Transform & Orchestrate",
                 "Tier 3 - Predictive/Spatial/Code", "Tier 4 - Self-Service/Interface"]:
        summary_data.append([tier, macro_tier_counts.get(tier, 0)])

    summary_data += [
        ["", ""],
        ["Tier Definitions", ""],
        ["Tier 1 - Simple ETL", "Simple data movement. Any platform works."],
        ["Tier 2 - Transform & Orchestrate", "Multi-step logic, macros, orchestration. ADF/Databricks/Python."],
        ["Tier 3 - Predictive/Spatial/Code", "R/Python/spatial/predictive tools. Databricks or Python."],
        ["Tier 4 - Self-Service/Interface", "User-facing interface elements. Needs product decision."],
        ["", ""],
        ["Unique Data Connections", len(set(c for r in results for c in r["connections"]))],
        ["Unique Macros Referenced", len(set(m for r in results for m in r["macros"]))],
        ["Workflows With Errors", sum(1 for r in results if r["errors"])],
    ]

    for row_idx, row_data in enumerate(summary_data, 1):
        for col_idx, val in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.border = thin_border
            if row_idx == 1:
                cell.font = Font(bold=True, size=14)
            elif row_idx == 5 or row_idx == 9:
                cell.font = header_font
                cell.fill = header_fill

    ws.column_dimensions["A"].width = 45
    ws.column_dimensions["B"].width = 60

    # ── Sheet 2: Workflow Inventory ───────────────────────────────────────
    ws2 = wb.create_sheet("Workflow Inventory")
    headers = ["Workflow Name", "Tier", "Tool Count", "Tools Used",
               "Data Connections", "Credentials", "Macros", "SQL Query Count",
               "Containers", "Version", "Errors"]

    for col_idx, h in enumerate(headers, 1):
        cell = ws2.cell(row=1, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border

    for row_idx, r in enumerate(results, 2):
        row_data = [
            r["name"],
            r["tier"],
            r["tool_count"],
            ", ".join(f"{k}({v})" for k, v in r["tools"].most_common()),
            "\n".join(r["connections"]) if r["connections"] else "",
            ", ".join(r["credentials"]) if r["credentials"] else "",
            "\n".join(r["macros"]) if r["macros"] else "",
            len(r["sql_queries"]),
            ", ".join(r["containers"]) if r["containers"] else "",
            r["version"],
            "; ".join(r["errors"]) if r["errors"] else "",
        ]
        for col_idx, val in enumerate(row_data, 1):
            cell = ws2.cell(row=row_idx, column=col_idx, value=val)
            cell.border = thin_border
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            if col_idx == 2 and val in tier_fills:
                cell.fill = tier_fills[val]

    ws2.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"
    for col_idx, w in enumerate([40, 35, 12, 60, 40, 20, 35, 14, 40, 10, 30], 1):
        ws2.column_dimensions[get_column_letter(col_idx)].width = w

    # ── Sheet 3: Tool Usage Across All Workflows ──────────────────────────
    ws3 = wb.create_sheet("Tool Usage")
    all_tools = Counter()
    for r in results:
        all_tools.update(r["tools"])

    tool_headers = ["Tool Name", "Total Occurrences", "Workflows Using It", "Category"]
    for col_idx, h in enumerate(tool_headers, 1):
        cell = ws3.cell(row=1, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border

    tool_wf_counts = Counter()
    for r in results:
        for t in r["tools"]:
            tool_wf_counts[t] += 1

    for row_idx, (tool, count) in enumerate(all_tools.most_common(), 2):
        cat = "Spatial" if tool in SPATIAL_TOOLS else \
              "Predictive/Code" if tool in PREDICTIVE_TOOLS else \
              "Interface" if tool in INTERFACE_TOOLS else \
              "Orchestration" if tool in ORCHESTRATION_TOOLS else "Standard"
        row_data = [tool, count, tool_wf_counts[tool], cat]
        for col_idx, val in enumerate(row_data, 1):
            cell = ws3.cell(row=row_idx, column=col_idx, value=val)
            cell.border = thin_border

    ws3.column_dimensions["A"].width = 30
    ws3.column_dimensions["B"].width = 20
    ws3.column_dimensions["C"].width = 22
    ws3.column_dimensions["D"].width = 20

    # ── Sheet 4: Data Connections ─────────────────────────────────────────
    ws4 = wb.create_sheet("Data Connections")
    conn_headers = ["Connection", "Used By Workflows"]
    for col_idx, h in enumerate(conn_headers, 1):
        cell = ws4.cell(row=1, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border

    conn_map = defaultdict(list)
    for r in results:
        for c in r["connections"]:
            conn_map[c].append(r["name"])

    for row_idx, (conn, wfs) in enumerate(sorted(conn_map.items()), 2):
        ws4.cell(row=row_idx, column=1, value=conn).border = thin_border
        cell = ws4.cell(row=row_idx, column=2, value=", ".join(wfs))
        cell.border = thin_border
        cell.alignment = Alignment(wrap_text=True)

    ws4.column_dimensions["A"].width = 60
    ws4.column_dimensions["B"].width = 80

    # ── Sheet 5: Macros ───────────────────────────────────────────────────
    ws5 = wb.create_sheet("Macros")
    macro_headers = ["Macro", "Used By Workflows"]
    for col_idx, h in enumerate(macro_headers, 1):
        cell = ws5.cell(row=1, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border

    macro_map = defaultdict(list)
    for r in results:
        for m in r["macros"]:
            macro_map[m].append(r["name"])

    for row_idx, (macro, wfs) in enumerate(sorted(macro_map.items()), 2):
        ws5.cell(row=row_idx, column=1, value=macro).border = thin_border
        cell = ws5.cell(row=row_idx, column=2, value=", ".join(wfs))
        cell.border = thin_border
        cell.alignment = Alignment(wrap_text=True)

    ws5.column_dimensions["A"].width = 50
    ws5.column_dimensions["B"].width = 80

    # ── Sheet 6: SQL Queries ──────────────────────────────────────────────
    ws6 = wb.create_sheet("SQL Queries")
    sql_headers = ["Workflow", "SQL Query"]
    for col_idx, h in enumerate(sql_headers, 1):
        cell = ws6.cell(row=1, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border

    row_idx = 2
    for r in results:
        for sql in r["sql_queries"]:
            ws6.cell(row=row_idx, column=1, value=r["name"]).border = thin_border
            cell = ws6.cell(row=row_idx, column=2, value=sql)
            cell.border = thin_border
            cell.alignment = Alignment(wrap_text=True)
            row_idx += 1

    ws6.column_dimensions["A"].width = 40
    ws6.column_dimensions["B"].width = 120

    # Save
    wb.save(output_path)
    print(f"\nReport saved: {output_path}")


def main():
    if len(sys.argv) < 2:
        print("Usage: python Analyse-AlteryxWorkflows.py <source_dir> [output.xlsx]")
        print("  source_dir: folder containing .yxzp or .yxmd files")
        sys.exit(1)

    source_dir = sys.argv[1]
    output_path = sys.argv[2] if len(sys.argv) > 2 else os.path.join(source_dir, "migration_report.xlsx")

    if not os.path.isdir(source_dir):
        print(f"ERROR: {source_dir} is not a directory")
        sys.exit(1)

    print(f"Scanning: {source_dir}")
    workflow_files, yxmc_files, temp_dirs = find_yxmd_files(source_dir)
    print(f"Found {len(workflow_files)} workflow files (.yxmd/.yxwz/.yxwg), {len(yxmc_files)} macros (.yxmc)")

    results = []
    for i, f in enumerate(workflow_files, 1):
        print(f"  [{i}/{len(workflow_files)}] Parsing: {f.name}")
        r = parse_workflow(f)
        results.append(r)

    # Also parse macros for the inventory
    for f in yxmc_files:
        r = parse_workflow(f, is_macro=True)
        r["name"] = f"[MACRO] {r['name']}"
        results.append(r)

    print(f"\nParsed {len(results)} workflows/macros total")

    wf_results = [r for r in results if not r.get("is_macro")]
    macro_results = [r for r in results if r.get("is_macro")]

    # Tier summary — workflows only
    tier_counts = Counter(r["tier"] for r in wf_results)
    print(f"\n── Workflow Tier Breakdown ({len(wf_results)} workflows) ──")
    for tier, count in sorted(tier_counts.items()):
        print(f"  {tier}: {count}")

    macro_tier_counts = Counter(r["tier"] for r in macro_results)
    print(f"\n── Macro Tier Breakdown ({len(macro_results)} macros) ──")
    for tier, count in sorted(macro_tier_counts.items()):
        print(f"  {tier}: {count}")

    build_report(results, output_path)

    # Cleanup temp dirs
    for d in temp_dirs:
        shutil.rmtree(d, ignore_errors=True)


if __name__ == "__main__":
    main()