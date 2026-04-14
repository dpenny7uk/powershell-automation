"""
Alteryx Workflow Ownership Chase List Builder
==============================================
Cross-references the scoped migration report to resolve as many unowned
workflow owners as possible, then produces an Excel chase list for the rest.

Usage:
    python Build-OwnershipChaseList.py <scoped_migration_report.xlsx>

Output:
    ownership_chase_list.xlsx (same directory as input)
"""

import sys, os, re
from collections import defaultdict
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side


def load_report(path):
    """Load all sheets from the scoped migration report."""
    wb = load_workbook(path, data_only=True)

    def sheet_to_dicts(name):
        ws = wb[name]
        headers = [cell.value for cell in ws[1]]
        rows = []
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
            rows.append(dict(zip(headers, row)))
        return rows

    return {
        "workflows": sheet_to_dicts("Active Workflows"),
        "schedules": sheet_to_dicts("Schedules"),
        "jobs": sheet_to_dicts("Job Activity"),
    }


# ── Naming pattern rules ────────────────────────────────────────────────────
# Each rule: (regex_pattern, team_name, suggested_contact)
# Ordered most specific first

NAMING_PATTERNS = [
    # Numbered reserving/Cognos pipeline (Daniel Tebbutt's domain)
    (r"^\d+\.(HICI?|HSA|HIB|HIG)_", "Reserving / Cognos Pipeline", "Daniel Tebbutt (Daniel.Tebbutt@HISCOX.com)"),
    (r"^\d+\.(Load|Transform|Format|Land)_", "Data Load Pipeline", None),
    (r"Cognos", "Cognos / BI", "Daniel Tebbutt (Daniel.Tebbutt@HISCOX.com)"),

    # French operations (Roméo Cros's domain)
    (r"Process\s?RN", "French Operations (RN Processing)", "Roméo Cros (Romeo.Cros@HISCOX.com)"),
    (r"Process\s?MED", "French Operations (MED Processing)", "Roméo Cros (Romeo.Cros@HISCOX.com)"),
    (r"MeF|MARSH", "French Operations", "Roméo Cros (Romeo.Cros@HISCOX.com)"),

    # Cyber
    (r"Cyber", "Cyber Insurance", None),

    # European entities
    (r"EU_IR|EU_FR|EU_NL|EU_BE|EU_", "European Entities", "Tiago Gimenez Jacinto (tiagogimenez.jacinto@hiscox.com)"),

    # EPBCS / Finance
    (r"EPBCS", "Finance / EPBCS", "Escher Luton (escher.luton@hiscox.com)"),
    (r"Re_Cube|FEPP|FDW", "Finance / Reserving", None),

    # Schemes / Flood
    (r"Scheme|BDX", "Schemes", "Mariana Cardoso (mariana.cardoso@hiscox.com)"),
    (r"[Ff]lood", "Flood Re", "Mariana Cardoso (mariana.cardoso@hiscox.com)"),
    (r"BIKMO", "Schemes (BIKMO)", "Mariana Cardoso (mariana.cardoso@hiscox.com)"),

    # Claims
    (r"[Cc]laim", "Claims", None),
    (r"LMDR", "Claims / LMDR", None),

    # Reporting
    (r"[Rr]eport|Dashboard|Tableau", "Reporting / Dashboards", None),

    # Recertification / IT
    (r"Recert|Notification|DPD", "IT / Recertification", None),

    # Alternative Risk
    (r"Alternative Risk|Lineage|Acc_Loc", "Alternative Risk", None),
]


def classify_by_name(workflow_name):
    """Match a workflow name against naming patterns."""
    for pattern, team, contact in NAMING_PATTERNS:
        if re.search(pattern, workflow_name):
            return team, contact
    return "Unknown", None


def resolve_owners(data):
    """
    Cross-reference all data sources to resolve ownership.

    Resolution priority:
    1. Created By from Job Activity (direct evidence — someone ran it)
    2. Name-pattern matching (inferred team/contact)
    3. Truly unknown
    """
    workflows = data["workflows"]
    jobs_by_id = {r["Workflow ID"]: r for r in data["jobs"]}

    results = []

    for wf in workflows:
        owner = wf.get("Owner")
        is_unowned = not owner or str(owner) == "None"

        if not is_unowned:
            continue

        # Skip historic workflows — only chase ownership for recent activity (2025+)
        activity = wf.get("Activity Status", "")
        if "Historic" in activity:
            continue

        wf_id = wf.get("Workflow ID", "")
        wf_name = wf.get("Workflow Name", "")
        status = wf.get("Activity Status", "")
        tier = wf.get("Tier", "")
        last_job = wf.get("Last Job Date", "")
        tool_count = wf.get("Tool Count", 0)

        # ── Resolution attempt 1: Created By from report ──
        created_by = wf.get("Created By") or ""
        creator_email = wf.get("Creator Email") or ""

        # ── Resolution attempt 2: Job Activity sheet ──
        if not created_by and wf_id in jobs_by_id:
            job = jobs_by_id[wf_id]
            created_by = job.get("Created By", "") or ""
            creator_email = job.get("Creator Email", "") or ""

        # ── Resolution attempt 3: Name pattern ──
        team, suggested_contact = classify_by_name(wf_name)

        # Determine resolution status
        if created_by:
            resolution = "Resolved - Job Activity"
            resolved_owner = f"{created_by} ({creator_email})" if creator_email else created_by
        elif suggested_contact:
            resolution = "Likely - Name Pattern"
            resolved_owner = suggested_contact
        elif team != "Unknown":
            resolution = "Team Identified - Needs Contact"
            resolved_owner = f"[{team} team - contact unknown]"
        else:
            resolution = "Unknown - Needs Investigation"
            resolved_owner = ""

        results.append({
            "workflow_name": wf_name,
            "workflow_id": wf_id,
            "activity_status": status,
            "tier": tier,
            "tool_count": tool_count,
            "last_job_date": last_job,
            "resolution": resolution,
            "resolved_owner": resolved_owner,
            "team": team,
            "created_by": created_by,
            "creator_email": creator_email,
            "suggested_contact": suggested_contact or "",
        })

    return results


def build_chase_list(results, output_path):
    """Generate the Excel chase list."""
    wb = Workbook()

    # Styles
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill("solid", fgColor="2F5496")
    resolved_fill = PatternFill("solid", fgColor="C6EFCE")
    likely_fill = PatternFill("solid", fgColor="FFEB9C")
    team_fill = PatternFill("solid", fgColor="D9E1F2")
    unknown_fill = PatternFill("solid", fgColor="FFC7CE")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )

    status_fills = {
        "Resolved - Job Activity": resolved_fill,
        "Likely - Name Pattern": likely_fill,
        "Team Identified - Needs Contact": team_fill,
        "Unknown - Needs Investigation": unknown_fill,
    }

    # ── Sheet 1: Summary ─────────────────────────────────────────────────
    ws = wb.active
    ws.title = "Summary"

    resolved = sum(1 for r in results if r["resolution"] == "Resolved - Job Activity")
    likely = sum(1 for r in results if r["resolution"] == "Likely - Name Pattern")
    team_only = sum(1 for r in results if r["resolution"] == "Team Identified - Needs Contact")
    unknown = sum(1 for r in results if r["resolution"] == "Unknown - Needs Investigation")

    summary_data = [
        ["Alteryx Ownership Chase List", ""],
        ["", ""],
        ["Total Unowned Workflows", len(results)],
        ["", ""],
        ["Resolution Status", "Count"],
        ["Resolved - Owner found via Job Activity", resolved],
        ["Likely - Matched by naming pattern", likely],
        ["Team Identified - Team known, contact unknown", team_only],
        ["Unknown - Needs manual investigation", unknown],
        ["", ""],
        ["Resolution Rate", f"{(resolved + likely) / len(results) * 100:.0f}% confident, {(resolved + likely + team_only) / len(results) * 100:.0f}% with team matches"],
        ["", ""],
        ["Colour Key", ""],
        ["Green", "Resolved — owner confirmed from job history"],
        ["Amber", "Likely — name pattern matched to a known contact"],
        ["Blue", "Team identified — team known but no specific contact"],
        ["Red", "Unknown — needs manual investigation"],
        ["", ""],
        ["Next Steps", ""],
        ["1. Verify green rows", "Confirm the job runner is the right owner (they may have run it for someone else)"],
        ["2. Contact amber rows", "Reach out to the suggested contact to confirm ownership"],
        ["3. Investigate blue rows", "Find the right person within the identified team"],
        ["4. Escalate red rows", "Check Alteryx Server MongoDB audit logs, or ask team leads"],
    ]

    for row_idx, row_data in enumerate(summary_data, 1):
        for col_idx, val in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.border = thin_border
            if row_idx == 1:
                cell.font = Font(bold=True, size=14)
            elif row_idx == 5:
                cell.font = header_font
                cell.fill = header_fill
            elif row_idx == 6:
                cell.fill = resolved_fill
            elif row_idx == 7:
                cell.fill = likely_fill
            elif row_idx == 8:
                cell.fill = team_fill
            elif row_idx == 9:
                cell.fill = unknown_fill

    ws.column_dimensions["A"].width = 45
    ws.column_dimensions["B"].width = 75

    # ── Sheet 2: Full Chase List ─────────────────────────────────────────
    ws2 = wb.create_sheet("Chase List")
    headers = [
        "Workflow Name", "Resolution Status", "Resolved Owner / Contact",
        "Team / Domain", "Activity Status", "Tier", "Tool Count",
        "Last Job Date", "Action Required"
    ]

    for col_idx, h in enumerate(headers, 1):
        cell = ws2.cell(row=1, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border

    # Sort: unknown first (need action), then team-only, then likely, then resolved
    priority_order = {
        "Unknown - Needs Investigation": 0,
        "Team Identified - Needs Contact": 1,
        "Likely - Name Pattern": 2,
        "Resolved - Job Activity": 3,
    }
    sorted_results = sorted(results, key=lambda r: priority_order.get(r["resolution"], 99))

    for row_idx, r in enumerate(sorted_results, 2):
        # Determine action needed
        if r["resolution"] == "Resolved - Job Activity":
            action = "Verify with the user that they own this workflow"
        elif r["resolution"] == "Likely - Name Pattern":
            action = f"Contact {r['suggested_contact']} to confirm ownership"
        elif r["resolution"] == "Team Identified - Needs Contact":
            action = f"Find the right contact in the {r['team']} team"
        else:
            action = "Check MongoDB audit logs or ask team leads"

        row_data = [
            r["workflow_name"],
            r["resolution"],
            r["resolved_owner"],
            r["team"],
            r["activity_status"],
            r["tier"],
            r["tool_count"],
            str(r["last_job_date"]) if r["last_job_date"] else "",
            action,
        ]

        for col_idx, val in enumerate(row_data, 1):
            cell = ws2.cell(row=row_idx, column=col_idx, value=val)
            cell.border = thin_border
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            if col_idx == 2 and val in status_fills:
                cell.fill = status_fills[val]

    ws2.auto_filter.ref = f"A1:I1"
    for col_idx, w in enumerate([40, 30, 40, 25, 30, 30, 12, 20, 50], 1):
        from openpyxl.utils import get_column_letter
        ws2.column_dimensions[get_column_letter(col_idx)].width = w

    # ── Sheet 3: By Team ─────────────────────────────────────────────────
    ws3 = wb.create_sheet("By Team")
    team_headers = ["Team / Domain", "Workflow Count", "Resolved", "Unresolved", "Workflows", "Suggested Contact"]

    for col_idx, h in enumerate(team_headers, 1):
        cell = ws3.cell(row=1, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border

    # Group by team
    teams = defaultdict(list)
    for r in results:
        teams[r["team"]].append(r)

    row_idx = 2
    for team in sorted(teams.keys()):
        members = teams[team]
        resolved_count = sum(1 for m in members if "Resolved" in m["resolution"] or "Likely" in m["resolution"])
        unresolved_count = len(members) - resolved_count
        wf_names = ", ".join(m["workflow_name"] for m in members)
        contacts = set(m["suggested_contact"] for m in members if m["suggested_contact"])
        contact_str = " / ".join(contacts) if contacts else "Unknown"

        row_data = [team, len(members), resolved_count, unresolved_count, wf_names, contact_str]
        for col_idx, val in enumerate(row_data, 1):
            cell = ws3.cell(row=row_idx, column=col_idx, value=val)
            cell.border = thin_border
            cell.alignment = Alignment(wrap_text=True, vertical="top")
        row_idx += 1

    from openpyxl.utils import get_column_letter
    for col_idx, w in enumerate([30, 15, 12, 14, 80, 45], 1):
        ws3.column_dimensions[get_column_letter(col_idx)].width = w

    # ── Sheet 4: Resolved Owners ─────────────────────────────────────────
    ws4 = wb.create_sheet("Resolved Owners")
    owner_headers = ["Resolved Owner", "Workflow Count", "Workflows", "Resolution Method"]

    for col_idx, h in enumerate(owner_headers, 1):
        cell = ws4.cell(row=1, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border

    # Group by resolved owner
    owners = defaultdict(list)
    for r in results:
        if r["resolved_owner"]:
            owners[r["resolved_owner"]].append(r)

    row_idx = 2
    for owner in sorted(owners.keys()):
        wfs = owners[owner]
        wf_names = ", ".join(w["workflow_name"] for w in wfs)
        methods = set(w["resolution"] for w in wfs)

        row_data = [owner, len(wfs), wf_names, " / ".join(methods)]
        for col_idx, val in enumerate(row_data, 1):
            cell = ws4.cell(row=row_idx, column=col_idx, value=val)
            cell.border = thin_border
            cell.alignment = Alignment(wrap_text=True, vertical="top")
        row_idx += 1

    for col_idx, w in enumerate([45, 15, 80, 30], 1):
        ws4.column_dimensions[get_column_letter(col_idx)].width = w

    wb.save(output_path)
    print(f"\nChase list saved: {output_path}")


def main():
    if len(sys.argv) < 2:
        print("Usage: python Build-OwnershipChaseList.py <scoped_migration_report.xlsx>")
        sys.exit(1)

    input_path = sys.argv[1]
    if not os.path.isfile(input_path):
        print(f"ERROR: {input_path} not found")
        sys.exit(1)

    output_path = sys.argv[2] if len(sys.argv) > 2 else os.path.join(os.path.dirname(input_path), "ownership_chase_list.xlsx")

    print(f"Loading: {input_path}")
    data = load_report(input_path)
    print(f"  Active Workflows: {len(data['workflows'])}")
    print(f"  Schedules: {len(data['schedules'])}")
    print(f"  Job Activity: {len(data['jobs'])}")

    print("\nResolving ownership...")
    results = resolve_owners(data)

    # Print summary
    resolved = sum(1 for r in results if r["resolution"] == "Resolved - Job Activity")
    likely = sum(1 for r in results if r["resolution"] == "Likely - Name Pattern")
    team_only = sum(1 for r in results if r["resolution"] == "Team Identified - Needs Contact")
    unknown = sum(1 for r in results if r["resolution"] == "Unknown - Needs Investigation")

    print(f"\n  Total unowned:                   {len(results)}")
    print(f"  Resolved (from job history):     {resolved}")
    print(f"  Likely (name pattern match):     {likely}")
    print(f"  Team identified (no contact):    {team_only}")
    print(f"  Unknown (needs investigation):   {unknown}")
    print(f"  Resolution rate:                 {(resolved + likely) / len(results) * 100:.0f}% confident")

    build_chase_list(results, output_path)


if __name__ == "__main__":
    main()
