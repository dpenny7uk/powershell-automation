"""
Alteryx Migration - Final Scoped Report
========================================
Cross-references schedules, metadata, and the migration analysis
to produce a final report scoped to only active workflows.

Usage:
    python Build-ScopedReport.py <export_directory>

Expects these files in the directory:
    - schedules.csv          (from Export-AlteryxSchedules.ps1)
    - workflow_metadata.csv  (from Get-AlteryxMetadata.ps1)
    - migration_report.xlsx  (from Analyse-AlteryxWorkflows.py)

Optional:
    - users.csv              (from Export-AlteryxUsers.ps1 — resolves owner names)
    - collections.csv        (from Export-AlteryxCollections.ps1 — adds team/business grouping)
    - job_triggers.csv       (from Get-AlteryxJobTriggers.ps1 — adds Triggerers sheet)
    - job_triggers_summary.csv (from Get-AlteryxJobTriggers.ps1 — adds trigger columns & stats)

Output:
    - scoped_migration_report.xlsx
"""

import sys, os, re
from collections import defaultdict
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# ── Naming pattern rules for ownership inference ─────────────────────────────
# (regex, team_name, suggested_contact)
NAMING_PATTERNS = [
    (r"^\d+\.(HICI?|HSA|HIB|HIG)_", "Reserving / Cognos Pipeline", "Daniel Tebbutt (Daniel.Tebbutt@HISCOX.com)"),
    (r"^\d+\.(Load|Transform|Format|Land)_", "Data Load Pipeline", None),
    (r"Cognos", "Cognos / BI", "Daniel Tebbutt (Daniel.Tebbutt@HISCOX.com)"),
    (r"Process\s?RN", "French Operations (RN Processing)", "Roméo Cros (Romeo.Cros@HISCOX.com)"),
    (r"Process\s?MED", "French Operations (MED Processing)", "Roméo Cros (Romeo.Cros@HISCOX.com)"),
    (r"MeF|MARSH", "French Operations", "Roméo Cros (Romeo.Cros@HISCOX.com)"),
    (r"Cyber", "Cyber Insurance", None),
    (r"EU_IR|EU_FR|EU_NL|EU_BE|EU_", "European Entities", "Tiago Gimenez Jacinto (tiagogimenez.jacinto@hiscox.com)"),
    (r"EPBCS", "Finance / EPBCS", "Escher Luton (escher.luton@hiscox.com)"),
    (r"Re_Cube|FEPP|FDW", "Finance / Reserving", None),
    (r"Scheme|BDX", "Schemes", "Mariana Cardoso (mariana.cardoso@hiscox.com)"),
    (r"[Ff]lood", "Flood Re", "Mariana Cardoso (mariana.cardoso@hiscox.com)"),
    (r"BIKMO", "Schemes (BIKMO)", "Mariana Cardoso (mariana.cardoso@hiscox.com)"),
    (r"[Cc]laim", "Claims", None),
    (r"LMDR", "Claims / LMDR", None),
    (r"[Rr]eport|Dashboard|Tableau", "Reporting / Dashboards", None),
    (r"Recert|Notification|DPD", "IT / Recertification", None),
    (r"Alternative Risk|Lineage|Acc_Loc", "Alternative Risk", None),
]


def infer_owner_by_name(workflow_name):
    """Match a workflow name against naming patterns to infer team/contact."""
    for pattern, team, contact in NAMING_PATTERNS:
        if re.search(pattern, workflow_name):
            return team, contact
    return "", None


def flag_tier_review(tier, is_scheduled):
    """
    Scheduled analytic apps usually run in batch mode with default answers and
    should migrate as Tier 2 (orchestration), but some may genuinely be
    user-initiated apps that happen to also have a schedule. Flag for human
    review rather than silently rewriting the tier.

    Returns (review_flag, review_reason).
    """
    if is_scheduled and tier == "Tier 4 - Self-Service/Interface":
        return (
            "Needs Manual Review",
            "Scheduled analytic app — likely Tier 2 (batch), confirm whether UI is actually used",
        )
    return ("", "")

def main():
    if len(sys.argv) < 2:
        print("Usage: python Build-ScopedReport.py <directory>")
        sys.exit(1)

    base = sys.argv[1]
    sched_path = os.path.join(base, "schedules.csv")
    meta_path = os.path.join(base, "workflow_metadata.csv")
    report_path = os.path.join(base, "migration_report.xlsx")
    users_path = os.path.join(base, "users.csv")
    triggers_path = os.path.join(base, "job_triggers.csv")
    triggers_summary_path = os.path.join(base, "job_triggers_summary.csv")
    output_path = os.path.join(base, "scoped_migration_report.xlsx")

    for f, label in [(sched_path, "schedules.csv"), (meta_path, "workflow_metadata.csv"), (report_path, "migration_report.xlsx")]:
        if not os.path.exists(f):
            print(f"ERROR: {label} not found in {base}")
            sys.exit(1)

    # ── Load data ─────────────────────────────────────────────────────────
    print("Loading schedules...")
    schedules = pd.read_csv(sched_path)

    print("Loading metadata...")
    metadata = pd.read_csv(meta_path)

    print("Loading migration report...")
    inventory = pd.read_excel(report_path, sheet_name="Workflow Inventory")
    tool_usage = pd.read_excel(report_path, sheet_name="Tool Usage")

    # ── Load users (optional) ─────────────────────────────────────────────
    user_lookup = {}
    if os.path.exists(users_path):
        print("Loading users...")
        users = pd.read_csv(users_path)
        for _, u in users.iterrows():
            uid = str(u.get("id", "")).strip()
            first = str(u.get("firstName", "")).strip()
            last = str(u.get("lastName", "")).strip()
            email = str(u.get("email", "")).strip()
            user_lookup[uid] = f"{first} {last} ({email})" if email and email != "nan" else f"{first} {last}"
        print(f"  Loaded {len(user_lookup)} users")
    else:
        print("  users.csv not found — owner names will not be resolved")

    # ── Load collections (optional) ───────────────────────────────────────
    collections_path = os.path.join(base, "collections.csv")
    coll_lookup = defaultdict(list)  # workflowId -> list of collection names
    if os.path.exists(collections_path):
        print("Loading collections...")
        collections = pd.read_csv(collections_path)
        for _, c in collections.iterrows():
            wf_id = str(c.get("workflowId", "")).strip()
            coll_name = str(c.get("collectionName", "")).strip()
            if wf_id and wf_id != "nan" and coll_name and coll_name != "nan":
                if coll_name not in coll_lookup[wf_id]:
                    coll_lookup[wf_id].append(coll_name)
        print(f"  Loaded {len(coll_lookup)} workflow-collection mappings")
    else:
        print("  collections.csv not found — collection names will not be included")

    # ── Load job triggers (optional) ─────────────────────────────────────
    trigger_lookup = {}       # workflowId -> summary row dict
    trigger_detail = None     # full DataFrame for Triggerers sheet
    if os.path.exists(triggers_summary_path):
        print("Loading job triggers summary...")
        trig_summary = pd.read_csv(triggers_summary_path)
        for _, t in trig_summary.iterrows():
            wf_id = str(t.get("workflowId", "")).strip()
            if wf_id and wf_id != "nan":
                trigger_lookup[wf_id] = {
                    "createdByName": str(t.get("createdByName", "")),
                    "createdByEmail": str(t.get("createdByEmail", "")),
                    "totalRuns": int(t.get("totalRuns", 0)),
                }
        print(f"  Loaded trigger summaries for {len(trigger_lookup)} workflows")
    else:
        print("  job_triggers_summary.csv not found — trigger data will not be included")

    if os.path.exists(triggers_path):
        print("Loading job triggers detail...")
        trigger_detail = pd.read_csv(triggers_path)
        print(f"  Loaded {len(trigger_detail)} job trigger records")
    else:
        print("  job_triggers.csv not found — Triggerers sheet will not be generated")

    # ── Get unique scheduled workflow IDs ─────────────────────────────────
    all_scheduled_ids = set(schedules["workflowId"].dropna().unique())
    print(f"Unique scheduled workflows (all): {len(all_scheduled_ids)}")

    # Filter to enabled schedules only if the column exists
    if "enabled" in schedules.columns:
        enabled_schedules = schedules[schedules["enabled"].astype(str).str.strip().str.upper().isin(["TRUE", "1"])]
        disabled_schedules = schedules[~schedules["enabled"].astype(str).str.strip().str.upper().isin(["TRUE", "1"])]
        scheduled_ids = set(enabled_schedules["workflowId"].dropna().unique())
        disabled_ids = set(disabled_schedules["workflowId"].dropna().unique()) - scheduled_ids
        print(f"Unique scheduled workflows (enabled): {len(scheduled_ids)}")
        print(f"Unique scheduled workflows (disabled only): {len(disabled_ids)}")
    else:
        scheduled_ids = all_scheduled_ids
        disabled_ids = set()
        print("  (no 'enabled' column found — treating all as enabled)")

    print(f"Using {len(scheduled_ids)} enabled scheduled workflows as primary scope")

    # ── Classify all workflows ────────────────────────────────────────────
    metadata["is_scheduled"] = metadata["id"].isin(scheduled_ids)

    # Convert runCount to numeric, coerce errors to 0
    metadata["runCount"] = pd.to_numeric(metadata["runCount"], errors="coerce").fillna(0).astype(int)

    # Parse lastJobDate
    metadata["lastJobDate_parsed"] = pd.to_datetime(metadata["lastJobDate"], errors="coerce", dayfirst=True)

    # Active but unscheduled: has run count > 0 OR has a job date, but no schedule
    metadata["is_active_unscheduled"] = (
        ~metadata["is_scheduled"] &
        (
            (metadata["runCount"] > 0) |
            (metadata["lastJobDate_parsed"].notna())
        )
    )

    # Parse dateCreated so we can spot newly-created workflows
    metadata["dateCreated_parsed"] = pd.to_datetime(metadata["dateCreated"], errors="coerce", dayfirst=True)

    # First-pass classification (coarse — refined after schedule merge)
    def classify_activity(row):
        if row["is_scheduled"]:
            return "Scheduled"
        elif row["is_active_unscheduled"]:
            return "Active - Unscheduled"
        else:
            return "Inactive"

    metadata["activity_status"] = metadata.apply(classify_activity, axis=1)

    # Scope: everything that isn't Inactive, PLUS newly-created 2026 workflows
    # (even if they have no runs yet — they're in scope for migration)
    in_scope_mask = (metadata["activity_status"] != "Inactive") | (
        metadata["dateCreated_parsed"].dt.year.fillna(0).astype(int) == 2026
    )
    active_meta = metadata[in_scope_mask].copy()

    # ── Schedule summary per workflow ─────────────────────────────────────
    sched_summary = schedules.groupby("workflowId").agg(
        schedule_count=("id", "count"),
        schedule_names=("name", lambda x: " | ".join(x.unique())),
        next_run=("runDateTime", "max")
    ).reset_index()

    active_meta = active_meta.merge(sched_summary, left_on="id", right_on="workflowId", how="left", suffixes=("", "_sched"))

    # ── Refine activity_status with schedule next-run and last-run year ──
    today = pd.Timestamp.now().normalize()

    def refine_activity_status(row):
        """
        Produces statuses like:
          - "Active - Scheduled 2026"      (schedule has a future run, last ran in 2026)
          - "Historic - Scheduled 2024"    (no future runs, last ran in 2024)
          - "Active - Unscheduled 2026"    (no schedule, recent manual runs)
          - "Historic - Unscheduled 2023"  (no schedule, old manual runs)
          - "New 2026 - Scheduled"         (created 2026, no runs yet, has a schedule)
          - "New 2026 - Unscheduled"       (created 2026, no runs yet, no schedule)
          - "Inactive"
        """
        last_run = row.get("lastJobDate_parsed")
        next_run = pd.to_datetime(row.get("next_run"), errors="coerce")
        date_created = row.get("dateCreated_parsed")
        year_suffix = f" {last_run.year}" if pd.notna(last_run) else ""

        # Newly-created 2026 workflow with no run history yet
        has_runs = pd.notna(last_run) or (row.get("runCount", 0) or 0) > 0
        is_new_2026 = pd.notna(date_created) and date_created.year == 2026 and not has_runs
        if is_new_2026:
            return "New 2026 - Scheduled" if row["is_scheduled"] else "New 2026 - Unscheduled"

        # Scheduled workflows: "Active" requires BOTH a future next-run AND a
        # recent actual job. A schedule with a future next-run but no real job
        # since 2024 is a stuck/broken schedule, not a live workflow.
        if row["is_scheduled"]:
            has_future_run = pd.notna(next_run) and next_run >= today
            has_recent_job = pd.notna(last_run) and last_run >= pd.Timestamp("2025-01-01")
            if has_future_run and has_recent_job:
                return f"Active - Scheduled{year_suffix}"
            return f"Historic - Scheduled{year_suffix}"

        # Unscheduled workflows with activity
        if row["is_active_unscheduled"]:
            if pd.notna(last_run) and last_run >= pd.Timestamp("2025-01-01"):
                return f"Active - Unscheduled{year_suffix}"
            return f"Historic - Unscheduled{year_suffix}"

        return "Inactive"

    active_meta["activity_status"] = active_meta.apply(refine_activity_status, axis=1)

    # Print refined breakdown
    print("\n── Activity Classification (refined) ──")
    for status, count in active_meta["activity_status"].value_counts().items():
        print(f"  {status}: {count}")

    print(f"\nTotal in scope (all active + new 2026): {len(active_meta)}")

    # ── Match to migration analysis by workflow name ──────────────────────
    # The migration report uses workflow names, metadata uses names too
    # Try to match on name
    inv_lookup = {}
    for _, row in inventory.iterrows():
        inv_lookup[str(row.get("Workflow Name", "")).strip().lower()] = row

    matched = []
    unmatched_names = []

    for _, row in active_meta.iterrows():
        wf_name = str(row.get("name", "")).strip()
        wf_name_lower = wf_name.lower()

        inv_row = inv_lookup.get(wf_name_lower)

        # Try partial match if exact fails
        if inv_row is None:
            for inv_name, inv_data in inv_lookup.items():
                if wf_name_lower in inv_name or inv_name in wf_name_lower:
                    inv_row = inv_data
                    break

        # Resolve owner — fallback chain:
        #   1. Schedule owner (from users.csv lookup)
        #   2. Created By (from job triggers — whoever last ran it)
        #   3. Name pattern inference (team/contact heuristic)
        owner_id = ""
        owner_name = ""
        owner_source = ""
        sched_rows = schedules[schedules["workflowId"] == row.get("id", "")]
        if len(sched_rows) > 0:
            owner_id = str(sched_rows.iloc[0].get("ownerId", "")).strip()
        if owner_id and owner_id in user_lookup:
            owner_name = user_lookup[owner_id]
            owner_source = "Schedule Owner"

        # Resolve collections
        wf_id = str(row.get("id", "")).strip()
        wf_collections = " | ".join(coll_lookup.get(wf_id, []))

        # Resolve trigger data
        trig = trigger_lookup.get(wf_id, {})

        # Fallback 2: Created By from job triggers
        created_by = trig.get("createdByName", "")
        creator_email = trig.get("createdByEmail", "")
        if created_by and created_by not in ("nan", "Unknown", ""):
            if not owner_name:
                owner_name = f"{created_by} ({creator_email})" if creator_email and creator_email != "nan" else created_by
                owner_source = "Job History"

        # Fallback 3: Name pattern inference
        team, suggested_contact = infer_owner_by_name(wf_name)
        if not owner_name and suggested_contact:
            owner_name = suggested_contact
            owner_source = "Name Pattern (unconfirmed)"
        elif not owner_name and team:
            owner_name = f"[{team} team]"
            owner_source = "Team Inference (unconfirmed)"

        matched.append({
            "Workflow Name": wf_name,
            "Workflow ID": row.get("id", ""),
            "Owner": owner_name,
            "Ownership Source": owner_source,
            "Collections": wf_collections,
            "Activity Status": row.get("activity_status", ""),
            "Date Created": row.get("dateCreated", ""),
            "Run Count": row.get("runCount", 0),
            "Package Type": row.get("packageType", ""),
            "Published": row.get("published", ""),
            "Run Disabled": row.get("runDisabled", ""),
            "Last Job Status": row.get("lastJobStatus", ""),
            "Last Job Date": row.get("lastJobDate", ""),
            "Schedule Count": row.get("schedule_count", 0) if pd.notna(row.get("schedule_count")) else 0,
            "Schedule Names": row.get("schedule_names", "") if pd.notna(row.get("schedule_names")) else "",
            "Next Scheduled Run": row.get("next_run", "") if pd.notna(row.get("next_run")) else "",
            "Created By": created_by if created_by not in ("nan", "Unknown", "") else "",
            "Creator Email": creator_email if creator_email not in ("nan", "") else "",
            "Total Runs": trig.get("totalRuns", ""),
            "Tier": inv_row["Tier"] if inv_row is not None else "Not in analysis",
            "Review Flag": flag_tier_review(
                inv_row["Tier"] if inv_row is not None else "",
                bool(row.get("is_scheduled", False)),
            )[0],
            "Review Reason": flag_tier_review(
                inv_row["Tier"] if inv_row is not None else "",
                bool(row.get("is_scheduled", False)),
            )[1],
            "Tool Count": inv_row["Tool Count"] if inv_row is not None else "",
            "Tools Used": inv_row["Tools Used"] if inv_row is not None else "",
            "Data Connections": inv_row["Data Connections"] if inv_row is not None else "",
            "Macros": inv_row["Macros"] if inv_row is not None else "",
            "SQL Query Count": inv_row["SQL Query Count"] if inv_row is not None else "",
        })

        if inv_row is None:
            unmatched_names.append(wf_name)

    df = pd.DataFrame(matched)
    print(f"\nTotal active workflows: {len(df)}")
    if unmatched_names:
        print(f"Workflows not matched to analysis: {len(unmatched_names)}")
        for n in unmatched_names[:10]:
            print(f"  - {n}")
        if len(unmatched_names) > 10:
            print(f"  ... and {len(unmatched_names) - 10} more")

    # ── Tier breakdown ────────────────────────────────────────────────────
    print("\n── Active Workflow Tier Breakdown ──")
    tier_counts = df["Tier"].value_counts()
    for tier, count in tier_counts.items():
        print(f"  {tier}: {count}")

    print("\n── By Activity Status ──")
    activity_counts = df["Activity Status"].value_counts()
    for status, count in activity_counts.items():
        print(f"  {status}: {count}")

    # ── Build Excel report ────────────────────────────────────────────────
    wb = Workbook()

    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill("solid", fgColor="2F5496")
    tier1_fill = PatternFill("solid", fgColor="C6EFCE")
    tier2_fill = PatternFill("solid", fgColor="FFEB9C")
    tier3_fill = PatternFill("solid", fgColor="FFC7CE")
    tier4_fill = PatternFill("solid", fgColor="D9E1F2")
    nomatch_fill = PatternFill("solid", fgColor="F2F2F2")
    scheduled_fill = PatternFill("solid", fgColor="C6EFCE")
    recent_fill = PatternFill("solid", fgColor="FFEB9C")
    historic_fill = PatternFill("solid", fgColor="FFC7CE")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )
    # Activity-status fills are matched by prefix because the new statuses
    # carry a year suffix (e.g. "Active - Scheduled 2026")
    def fill_for_activity_status(val):
        if not val:
            return None
        if val.startswith("Active - Scheduled"):
            return scheduled_fill   # green — top priority
        if val.startswith("Historic - Scheduled"):
            return recent_fill      # amber — was scheduled, now stale
        if val.startswith("Active - Unscheduled"):
            return recent_fill      # amber — recent manual usage
        if val.startswith("Historic - Unscheduled"):
            return historic_fill    # red — old manual usage
        if val.startswith("New 2026"):
            return scheduled_fill   # green — fresh, in scope
        return None
    tier_fills = {
        "Tier 1 - Simple ETL": tier1_fill,
        "Tier 2 - Transform & Orchestrate": tier2_fill,
        "Tier 3 - Predictive/Spatial/Code": tier3_fill,
        "Tier 4 - Self-Service/Interface": tier4_fill,
        "Not in analysis": nomatch_fill,
    }

    # ── Sheet 1: Executive Summary ────────────────────────────────────────
    ws = wb.active
    ws.title = "Summary"

    summary_rows = [
        ["Alteryx Migration - Scoped Active Workflows", ""],
        ["", ""],
        ["What This Report Shows", ""],
        ["This report is a full inventory of all workflows on the Alteryx Server,", ""],
        ["classified by usage to support migration planning.", ""],
        ["", ""],
        ["Scheduled = workflows running automatically on a timer (must migrate).", ""],
        ["Unscheduled = workflows run manually by users or via the API.", ""],
        ["Inactive = workflows never run or not run in a long time with no schedule.", ""],
        ["", ""],
        ["Tiers describe workflow complexity and determine which platform to migrate to.", ""],
        ["In Scope = everything that is actively used (Scheduled + Unscheduled).", ""],
        ["", ""],
        ["Action needed: assign owners to unscheduled workflows, agree target", ""],
        ["platform per tier, and prioritise scheduled workflows for migration.", ""],
        ["", ""],
        ["Total Workflows in Gallery", len(metadata)],
        ["", ""],
        ["Activity Classification", "Count"],
        ["Active - Scheduled (next run in future — HIGHEST PRIORITY)",
            int(active_meta["activity_status"].str.startswith("Active - Scheduled").sum())],
        ["Historic - Scheduled (schedule exists but no future runs)",
            int(active_meta["activity_status"].str.startswith("Historic - Scheduled").sum())],
        ["Active - Unscheduled (run manually since 2025)",
            int(active_meta["activity_status"].str.startswith("Active - Unscheduled").sum())],
        ["Historic - Unscheduled (ran manually, but not since 2025)",
            int(active_meta["activity_status"].str.startswith("Historic - Unscheduled").sum())],
        ["New 2026 - Scheduled (created this year, on a schedule)",
            int(active_meta["activity_status"].eq("New 2026 - Scheduled").sum())],
        ["New 2026 - Unscheduled (created this year, no schedule)",
            int(active_meta["activity_status"].eq("New 2026 - Unscheduled").sum())],
        ["Inactive (no runs, no jobs, no schedule, not new)",
            int(metadata["activity_status"].eq("Inactive").sum()) - int(active_meta["activity_status"].str.startswith("New 2026").sum())],
        ["", ""],
        ["Note: disabled schedules", int(len(disabled_ids))],
        ["(These workflows have schedules that are turned off. They are counted", ""],
        ["in the Unscheduled or Inactive categories above, not as Scheduled.)", ""],
        ["", ""],
        ["Total In Scope (all active)", len(df)],
        ["Workflows Matched to Analysis", len(df) - len(unmatched_names)],
        ["Workflows Not Matched (missing .yxzp)", len(unmatched_names)],
        ["", ""],
        ["Active Workflow Tier Breakdown", "Count"],
    ]
    for tier in ["Tier 1 - Simple ETL", "Tier 2 - Transform & Orchestrate",
                 "Tier 3 - Predictive/Spatial/Code", "Tier 4 - Self-Service/Interface",
                 "Not in analysis"]:
        summary_rows.append([tier, tier_counts.get(tier, 0)])

    # Count workflows flagged for manual tier review
    flagged_count = sum(1 for r in matched if r.get("Review Flag"))

    summary_rows += [
        ["", ""],
        ["Workflows Needing Manual Tier Review", "Count"],
        ["Scheduled Tier 4 (likely batch, not self-service)", flagged_count],
        ["", ""],
        ["Platform Recommendation", ""],
        ["Tier 1 - Simple ETL", "Any platform: ADF, Python, Power BI Dataflows, Databricks"],
        ["Tier 2 - Transform & Orchestrate", "ADF or Databricks (pipeline orchestration needed)"],
        ["Tier 3 - Predictive/Spatial/Code", "Databricks or standalone Python"],
        ["Tier 4 - Self-Service/Interface", "Needs product decision - user-facing capability"],
        ["Not in analysis", "Review manually - .yxzp not in export or name mismatch"],
    ]

    # Rows that should get header styling (dark blue background, white text)
    header_labels = {
        "Activity Classification", "Active Workflow Tier Breakdown", "Platform Recommendation",
        "What This Report Shows", "Workflows Needing Manual Tier Review",
    }

    for row_idx, row_data in enumerate(summary_rows, 1):
        for col_idx, val in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.border = thin_border
            if row_idx == 1:
                cell.font = Font(bold=True, size=14)
            elif row_data[0] in header_labels:
                cell.font = header_font
                cell.fill = header_fill

    ws.column_dimensions["A"].width = 50
    ws.column_dimensions["B"].width = 65

    # ── Sheet 2: Active Workflow Detail ───────────────────────────────────
    ws2 = wb.create_sheet("Active Workflows")
    headers = list(df.columns)

    for col_idx, h in enumerate(headers, 1):
        cell = ws2.cell(row=1, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border

    for row_idx, (_, row) in enumerate(df.iterrows(), 2):
        for col_idx, h in enumerate(headers, 1):
            val = row[h]
            if pd.isna(val):
                val = ""
            cell = ws2.cell(row=row_idx, column=col_idx, value=val)
            cell.border = thin_border
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            if h == "Tier" and val in tier_fills:
                cell.fill = tier_fills[val]
            if h == "Activity Status":
                fill = fill_for_activity_status(val)
                if fill is not None:
                    cell.fill = fill
            if h == "Ownership Source":
                source_fills = {
                    "Schedule Owner": scheduled_fill,
                    "Job History": tier2_fill,
                    "Name Pattern (unconfirmed)": tier4_fill,
                    "Team Inference (unconfirmed)": tier4_fill,
                }
                if val in source_fills:
                    cell.fill = source_fills[val]
                elif not val and not row.get("Owner", ""):
                    cell.fill = tier3_fill  # red — no owner at all
            if h == "Review Flag" and val == "Needs Manual Review":
                cell.fill = tier2_fill  # amber — needs attention

    ws2.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"
    col_widths = [35, 28, 35, 25, 40, 30, 18, 12, 14, 10, 12, 14, 18, 14, 40, 20, 40, 12, 30, 32, 20, 50, 12, 50, 35, 30, 14]
    for col_idx, w in enumerate(col_widths[:len(headers)], 1):
        ws2.column_dimensions[get_column_letter(col_idx)].width = w

    # ── Sheet 3: Unmatched Workflows ──────────────────────────────────────
    if unmatched_names:
        ws3 = wb.create_sheet("Unmatched")
        ws3.cell(row=1, column=1, value="Workflow Name").font = header_font
        ws3.cell(row=1, column=1).fill = header_fill
        ws3.cell(row=1, column=2, value="Notes").font = header_font
        ws3.cell(row=1, column=2).fill = header_fill
        for row_idx, name in enumerate(unmatched_names, 2):
            ws3.cell(row=row_idx, column=1, value=name).border = thin_border
            ws3.cell(row=row_idx, column=2, value="Not found in .yxzp export - review manually").border = thin_border
        ws3.column_dimensions["A"].width = 50
        ws3.column_dimensions["B"].width = 50

    # ── Sheet 4: Schedule Detail ──────────────────────────────────────────
    ws4 = wb.create_sheet("Schedules")
    sched_headers = list(schedules.columns)
    for col_idx, h in enumerate(sched_headers, 1):
        cell = ws4.cell(row=1, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border

    for row_idx, (_, row) in enumerate(schedules.iterrows(), 2):
        for col_idx, h in enumerate(sched_headers, 1):
            val = row[h]
            if pd.isna(val):
                val = ""
            cell = ws4.cell(row=row_idx, column=col_idx, value=val)
            cell.border = thin_border

    for col_idx in range(1, len(sched_headers) + 1):
        ws4.column_dimensions[get_column_letter(col_idx)].width = 30

    # ── Sheet 5: Job Activity (if job_triggers.csv is available) ────────────
    if trigger_detail is not None and len(trigger_detail) > 0:
        ws5 = wb.create_sheet("Job Activity")

        # Build per-workflow aggregation with creator info
        trig_agg = trigger_detail.groupby(["workflowId", "workflowName", "createdByName", "createdByEmail"]).agg(
            runCount=("jobId", "count"),
            firstRun=("jobDate", "min"),
            lastRun=("jobDate", "max"),
        ).reset_index()
        trig_agg = trig_agg.sort_values(["workflowName", "runCount"], ascending=[True, False])
        trig_agg = trig_agg.rename(columns={
            "workflowName": "Workflow Name",
            "workflowId": "Workflow ID",
            "createdByName": "Created By",
            "createdByEmail": "Creator Email",
            "runCount": "Run Count",
            "firstRun": "First Run",
            "lastRun": "Last Run",
        })

        trig_headers = list(trig_agg.columns)
        for col_idx, h in enumerate(trig_headers, 1):
            cell = ws5.cell(row=1, column=col_idx, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = thin_border

        for row_idx, (_, row) in enumerate(trig_agg.iterrows(), 2):
            for col_idx, h in enumerate(trig_headers, 1):
                val = row[h]
                if pd.isna(val):
                    val = ""
                cell = ws5.cell(row=row_idx, column=col_idx, value=val)
                cell.border = thin_border

        trig_col_widths = [35, 28, 30, 35, 12, 20, 20]
        for col_idx, w in enumerate(trig_col_widths[:len(trig_headers)], 1):
            ws5.column_dimensions[get_column_letter(col_idx)].width = w

        ws5.auto_filter.ref = f"A1:{get_column_letter(len(trig_headers))}1"

    # ── Add ownership stats to summary sheet ────────────────────────────────
    ws = wb["Summary"]
    last_row = ws.max_row + 2

    # Count by ownership source
    from collections import Counter
    source_counts = Counter(r.get("Ownership Source", "") for r in matched)
    has_owner = sum(1 for r in matched if r.get("Owner", ""))
    no_owner = sum(1 for r in matched if not r.get("Owner", ""))

    ownership_stats = [
        ["Ownership Resolution", ""],
        ["Total active workflows", len(matched)],
        ["Workflows with owner assigned", has_owner],
        ["Workflows with no owner", no_owner],
        ["", ""],
        ["Ownership Source Breakdown", "Count"],
        ["Schedule Owner (from Alteryx API)", source_counts.get("Schedule Owner", 0)],
        ["Job History (whoever last ran it)", source_counts.get("Job History", 0)],
        ["Name Pattern (inferred, unconfirmed)", source_counts.get("Name Pattern (unconfirmed)", 0)],
        ["Team Inference (team known, no contact)", source_counts.get("Team Inference (unconfirmed)", 0)],
        ["Unresolved", no_owner],
    ]

    for i, row_data in enumerate(ownership_stats):
        for col_idx, val in enumerate(row_data, 1):
            cell = ws.cell(row=last_row + i, column=col_idx, value=val)
            cell.border = thin_border
            if i == 0 or i == 5:
                cell.font = header_font
                cell.fill = header_fill

    wb.save(output_path)
    print(f"\nScoped report saved: {output_path}")


if __name__ == "__main__":
    main()