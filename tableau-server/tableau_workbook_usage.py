"""
Tableau Workbook Usage Report (Hybrid: Repository + REST API)
=============================================================
Queries Tableau Server's PostgreSQL repository (readonly) for usage stats,
optionally enriched with REST API data for project hierarchy, datasource
connection details, and workbook-datasource linkage.

Repository provides: workbook/view usage, staleness, flat project names.
REST API provides: project hierarchy, connection details, datasource linkage.

Outputs a dated Excel workbook with Summary, Workbook Usage, View Usage,
Datasources, and Projects sheets.

Environment Variables:
    TABLEAU_REPO_HOST       - Tableau Server hostname for repo (required)
    TABLEAU_REPO_PORT       - PostgreSQL port (default: 8060)
    TABLEAU_REPO_USER       - PostgreSQL readonly username (required)
    TABLEAU_REPO_PASSWORD   - Password (required)
    TABLEAU_PAT_NAME        - Personal Access Token name (optional, enables API)
    TABLEAU_PAT_SECRET      - Personal Access Token secret (optional)
    TABLEAU_SERVER_URL      - Server URL for REST API (optional)
    TABLEAU_API_VERSION     - REST API version (default: 3.25)
    TABLEAU_OUTPUT_DIR      - Output directory (default: ./output)
    TABLEAU_SITES           - Comma-separated site contentUrls (default: Default)
"""

import os
import sys
import logging
import shutil
import time
from datetime import datetime, timezone
from typing import Optional

from sqlalchemy import create_engine, text, bindparam
from sqlalchemy.engine.url import URL
import pandas as pd
import requests
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ============================================================================
# Configuration
# ============================================================================

# ============================================================================
# Configuration — all environment-driven, no hardcoded infrastructure values
# ============================================================================

TABLEAU_REPO_HOST = os.environ["TABLEAU_REPO_HOST"]  # Required
TABLEAU_REPO_PORT = int(os.getenv("TABLEAU_REPO_PORT", "8060"))
TABLEAU_REPO_DB = os.getenv("TABLEAU_REPO_DB", "workgroup")
TABLEAU_SERVER_URL = os.getenv("TABLEAU_SERVER_URL", f"http://{TABLEAU_REPO_HOST}")
TABLEAU_API_VERSION = os.getenv("TABLEAU_API_VERSION", "3.25")

OUTPUT_DIR = os.getenv("TABLEAU_OUTPUT_DIR", os.path.join(os.getcwd(), "output"))
SITES_FILTER = tuple(s.strip() for s in os.getenv("TABLEAU_SITES", "Default").split(","))

STALENESS_BINS = [-1, 30, 90, 180, 365, float("inf")]
STALENESS_LABELS = ["Active", "Slowing", "Stale", "Dormant", "Archive"]
MAX_HIERARCHY_DEPTH = 20  # Cycle detection guard for project hierarchy

# ============================================================================
# Logging
# ============================================================================

def setup_logging(output_dir: str, run_timestamp: datetime) -> logging.Logger:
    logger = logging.getLogger("tableau_usage")
    logger.setLevel(logging.INFO)
    if logger.handlers:
        return logger

    fmt = logging.Formatter("%(asctime)s [%(levelname)s] %(message)s", "%Y-%m-%d %H:%M:%S")

    console = logging.StreamHandler(sys.stdout)
    console.setFormatter(fmt)
    logger.addHandler(console)

    log_file = os.path.join(output_dir, f"tableau_usage_{run_timestamp.strftime('%Y%m%d_%H%M%S')}.log")
    fh = logging.FileHandler(log_file, encoding="utf-8")
    fh.setFormatter(fmt)
    logger.addHandler(fh)

    return logger

# ============================================================================
# SQL Queries — readonly views only
# ============================================================================
# _views_stats uses 'views_id' (not 'view_id').
# _data_connections does not exist as a readonly view.

WORKBOOK_QUERY = """
SELECT
    s.name                              AS site_name,
    s.url_namespace                     AS site_content_url,
    s.name || ' - ' || w.name          AS sitename_workbookname,
    w.name                              AS workbook_name,
    p.name                              AS project_name,
    su.name                             AS owner_username,
    su.friendly_name                    AS owner_display_name,
    w.created_at                        AS workbook_created_at,
    w.updated_at                        AS workbook_updated_at,
    COUNT(DISTINCT v.id)                AS view_count,
    MAX(vs.last_view_time)              AS workbook_last_accessed,
    COALESCE(SUM(vs.nviews), 0)         AS total_views_all_time
FROM _workbooks w
INNER JOIN _sites s          ON w.site_id = s.id
LEFT JOIN _projects p        ON w.project_id = p.id AND w.site_id = p.site_id
LEFT JOIN _system_users su   ON w.owner_id = su.id
LEFT JOIN _views v           ON v.workbook_id = w.id
LEFT JOIN _views_stats vs    ON vs.views_id = v.id
WHERE s.url_namespace IN :sites
GROUP BY s.name, s.url_namespace, w.id, w.name, p.name,
         su.name, su.friendly_name, w.created_at, w.updated_at
ORDER BY s.name, p.name, workbook_last_accessed ASC NULLS FIRST;
"""

VIEW_QUERY = """
SELECT
    s.name                              AS site_name,
    s.url_namespace                     AS site_content_url,
    w.name                              AS workbook_name,
    p.name                              AS project_name,
    su.name                             AS owner_username,
    su.friendly_name                    AS owner_display_name,
    v.name                              AS view_name,
    v.title                             AS view_title,
    COALESCE(vs.nviews, 0)              AS total_views_all_time,
    vs.last_view_time                   AS view_last_accessed
FROM _views v
INNER JOIN _workbooks w      ON v.workbook_id = w.id
INNER JOIN _sites s          ON w.site_id = s.id
LEFT JOIN _projects p        ON w.project_id = p.id AND w.site_id = p.site_id
LEFT JOIN _system_users su   ON w.owner_id = su.id
LEFT JOIN _views_stats vs    ON vs.views_id = v.id
WHERE s.url_namespace IN :sites
ORDER BY s.name, p.name, w.name, v.name;
"""

PROJECT_QUERY = """
SELECT
    s.name                              AS site_name,
    s.url_namespace                     AS site_content_url,
    p.name                              AS project_name,
    p.owner_name                        AS project_owner,
    COUNT(DISTINCT w.id)                AS workbook_count,
    p.created_at                        AS project_created_at
FROM _projects p
INNER JOIN _sites s    ON p.site_id = s.id
LEFT JOIN _workbooks w ON w.project_id = p.id AND w.site_id = p.site_id
WHERE s.url_namespace IN :sites
GROUP BY s.name, s.url_namespace, p.name, p.owner_name, p.created_at
ORDER BY s.name, p.name;
"""

DATASOURCE_QUERY = """
SELECT
    s.name                              AS site_name,
    s.url_namespace                     AS site_content_url,
    d.name                              AS datasource_name,
    su.name                             AS datasource_owner,
    su.friendly_name                    AS datasource_owner_display,
    d.created_at                        AS datasource_created_at,
    d.updated_at                        AS datasource_updated_at
FROM _datasources d
INNER JOIN _sites s         ON d.site_id = s.id
LEFT JOIN _system_users su  ON d.owner_id = su.id
WHERE s.url_namespace IN :sites
ORDER BY s.name, d.name;
"""

# ============================================================================
# REST API
# ============================================================================

class TableauRestAPI:
    """Thin wrapper around Tableau Server REST API."""

    def __init__(self, server_url: str, api_version: str, logger: logging.Logger):
        self.base_url = f"{server_url.rstrip('/')}/api/{api_version}"
        self.logger = logger
        self.session = requests.Session()
        self.session.headers.update({"Accept": "application/json", "Content-Type": "application/json"})
        self.auth_token: Optional[str] = None
        self.site_id: Optional[str] = None

    def sign_in(self, site_content_url: str, pat_name: str, pat_secret: str) -> bool:
        payload = {
            "credentials": {
                "personalAccessTokenName": pat_name,
                "personalAccessTokenSecret": pat_secret,
                "site": {"contentUrl": site_content_url},
            }
        }
        try:
            resp = self.session.post(f"{self.base_url}/auth/signin", json=payload, timeout=30)
            resp.raise_for_status()
            creds = resp.json()["credentials"]
            token: str = creds["token"]
            self.auth_token = token
            self.site_id = creds["site"]["id"]
            self.session.headers["X-Tableau-Auth"] = token
            return True
        except Exception as e:
            self.logger.error(f"API sign-in failed for site '{site_content_url}': {e}")
            return False

    def close(self):
        """Close the underlying HTTP session and release connections."""
        self.session.close()

    def sign_out(self):
        if self.auth_token:
            try:
                self.session.post(f"{self.base_url}/auth/signout", timeout=10)
            except Exception:
                pass
        self.auth_token = None
        self.site_id = None
        self.session.headers.pop("X-Tableau-Auth", None)

    def _get_paginated(self, endpoint: str, root_key: str, item_key: str) -> list:
        all_items = []
        page_number = 1
        page_size = 100
        expected_total = None

        while True:
            sep = "&" if "?" in endpoint else "?"
            url = f"{self.base_url}/sites/{self.site_id}/{endpoint}{sep}pageSize={page_size}&pageNumber={page_number}"
            try:
                resp = self.session.get(url, timeout=60)
                resp.raise_for_status()
                data = resp.json()
            except Exception as e:
                self.logger.error(f"API request failed: {endpoint} page {page_number}: {e}")
                if expected_total is not None:
                    self.logger.warning(
                        f"API pagination incomplete for {endpoint}: "
                        f"retrieved {len(all_items)} of {expected_total} items"
                    )
                break

            items = data.get(root_key, {}).get(item_key, [])
            if isinstance(items, dict):
                items = [items]
            all_items.extend(items)

            total = int(data.get("pagination", {}).get("totalAvailable", 0))
            if expected_total is None:
                expected_total = total
            if len(all_items) >= total or not items:
                break
            page_number += 1

        return all_items

    def get_projects(self) -> list:
        return self._get_paginated("projects", "projects", "project")

    def get_workbooks(self) -> list:
        return self._get_paginated("workbooks", "workbooks", "workbook")

    def get_datasources(self) -> list:
        return self._get_paginated("datasources", "datasources", "datasource")

    def get_workbook_connections(self, workbook_id: str) -> list:
        url = f"{self.base_url}/sites/{self.site_id}/workbooks/{workbook_id}/connections"
        try:
            resp = self.session.get(url, timeout=30)
            resp.raise_for_status()
            conns = resp.json().get("connections", {}).get("connection", [])
            return [conns] if isinstance(conns, dict) else conns
        except Exception as e:
            self.logger.warning(f"Failed to get connections for workbook {workbook_id}: {e}")
            return []

    def get_project_by_id(self, project_id: str) -> Optional[dict]:
        """Fetch a single project by ID — used for missing parent resolution."""
        url = f"{self.base_url}/sites/{self.site_id}/projects/{project_id}"
        try:
            resp = self.session.get(url, timeout=15)
            resp.raise_for_status()
            return resp.json().get("project", None)
        except Exception:
            return None


def build_project_hierarchy(projects: list, api: Optional["TableauRestAPI"] = None,
                            logger: Optional[logging.Logger] = None) -> dict:
    """Build project hierarchy lookup from API project list.

    When a parentProjectId isn't in the initial project list (user lacks
    direct access to the parent), attempts to fetch it via a direct API call.
    This resolves truncated hierarchies where intermediate or top-level
    projects are invisible to the PAT user's project listing.

    Returns dict keyed by project ID (not name) to handle duplicate names.
    Values: {name, top_level_project, project_path, project_depth}
    """
    by_id: dict[str, dict] = {p["id"]: p for p in projects}
    fetched_missing: set[str] = set()  # Track IDs we've already tried to fetch
    hierarchy: dict[str, dict] = {}

    for proj in projects:
        chain = [proj["name"]]
        current = proj
        visited = {proj["id"]}
        depth = 1

        while current.get("parentProjectId"):
            parent_id = current["parentProjectId"]
            if parent_id in visited:
                break
            if depth >= MAX_HIERARCHY_DEPTH:
                break

            # If parent not in our list, try fetching it directly
            if parent_id not in by_id and api is not None and parent_id not in fetched_missing:
                fetched_missing.add(parent_id)
                parent_proj = api.get_project_by_id(parent_id)
                if parent_proj and "id" in parent_proj:
                    by_id[parent_proj["id"]] = parent_proj
                    if logger:
                        logger.info(f"  Resolved missing parent project: '{parent_proj.get('name', '?')}'")

            if parent_id not in by_id:
                break

            visited.add(parent_id)
            parent = by_id[parent_id]
            chain.insert(0, parent["name"])
            current = parent
            depth += 1

        hierarchy[proj["id"]] = {
            "name": proj["name"],
            "top_level_project": chain[0],
            "project_path": "/".join(chain),
            "project_depth": depth,
        }

    return hierarchy


def fetch_api_data(
    server_url: str,
    api_version: str,
    pat_name: str,
    pat_secret: str,
    sites: tuple,
    logger: logging.Logger,
) -> dict:
    """Fetch project hierarchy and workbook connections from REST API."""
    api = TableauRestAPI(server_url, api_version, logger)
    all_hierarchies: dict[str, dict] = {}
    all_connections = []
    all_datasource_types = []
    all_workbook_project_map = []  # Maps workbook names to API project IDs

    try:
        for site_url in sites:
            logger.info(f"API: Signing in to site '{site_url}'...")
            signed_in = api.sign_in(site_url, pat_name, pat_secret)

            # Tableau's Default site uses contentUrl="" for API auth, but may have
            # a non-empty url_namespace in the repo. Retry with "" if first attempt fails.
            if not signed_in and site_url:
                logger.info(f"API: Retrying site '{site_url}' as Default site (contentUrl='')...")
                signed_in = api.sign_in("", pat_name, pat_secret)
                if signed_in:
                    # Validate we actually reached the intended site — a "" retry
                    # could match any Default site, not necessarily the one we wanted
                    try:
                        check_url = f"{api.base_url}/sites/{api.site_id}"
                        resp = api.session.get(check_url, timeout=15)
                        resp.raise_for_status()
                        actual_name = resp.json().get("site", {}).get("name", "")
                        if actual_name.lower() != site_url.lower():
                            logger.warning(
                                f"API: Default site retry returned '{actual_name}' "
                                f"but expected '{site_url}' — skipping to avoid data mismatch."
                            )
                            api.sign_out()
                            signed_in = False
                    except Exception:
                        logger.warning(f"API: Could not validate site identity after Default retry — skipping '{site_url}'.")
                        api.sign_out()
                        signed_in = False

            if not signed_in:
                logger.warning(f"API: Skipping site '{site_url}' — sign-in failed.")
                continue

            try:
                # Projects
                projects = api.get_projects()
                logger.info(f"API [{site_url}]: {len(projects)} projects")
                hierarchy = build_project_hierarchy(projects, api=api, logger=logger)
                all_hierarchies[site_url] = hierarchy  # ID-keyed — merge via workbook project map

                # Datasources
                datasources = api.get_datasources()
                logger.info(f"API [{site_url}]: {len(datasources)} datasources")
                for ds in datasources:
                    all_datasource_types.append({
                        "site_content_url": site_url,
                        "datasource_name": ds.get("name", ""),
                        "datasource_type": ds.get("type", ""),
                        "datasource_has_extracts": ds.get("hasExtracts", False),
                    })

                # Workbook connections
                workbooks = api.get_workbooks()
                logger.info(f"API [{site_url}]: {len(workbooks)} workbooks — fetching connections...")

                for i, wb in enumerate(workbooks, 1):
                    # Map workbook to its API project ID for accurate hierarchy merge
                    wb_project_id = wb.get("project", {}).get("id", "")
                    if wb_project_id:
                        all_workbook_project_map.append({
                            "site_content_url": site_url,
                            "workbook_name": wb.get("name", ""),
                            "project_name": wb.get("project", {}).get("name", ""),
                            "api_project_id": wb_project_id,
                        })

                    for conn in api.get_workbook_connections(wb["id"]):
                        ds_info = conn.get("datasource", {})
                        all_connections.append({
                            "site_content_url": site_url,
                            "workbook_name": wb.get("name", ""),
                            "project_name": wb.get("project", {}).get("name", ""),
                            "connection_type": conn.get("type", ""),
                            "connection_server": conn.get("serverAddress", ""),
                            "connection_port": conn.get("serverPort", ""),
                            "connection_username": conn.get("userName", ""),
                            "connection_database": conn.get("dbName", ""),
                            "datasource_name": ds_info.get("name", ""),
                        })
                    if i % 50 == 0:
                        logger.info(f"API [{site_url}]: {i}/{len(workbooks)} workbooks processed...")
                    time.sleep(0.05)

                site_conns = sum(1 for c in all_connections if c["site_content_url"] == site_url)
                logger.info(f"API [{site_url}]: complete — {site_conns} connections")
            finally:
                api.sign_out()
    finally:
        api.close()

    return {
        "project_hierarchy": all_hierarchies,
        "workbook_project_map": pd.DataFrame(all_workbook_project_map) if all_workbook_project_map else pd.DataFrame(),
        "workbook_connections": pd.DataFrame(all_connections) if all_connections else pd.DataFrame(),
        "datasource_types": pd.DataFrame(all_datasource_types) if all_datasource_types else pd.DataFrame(),
    }

# ============================================================================
# Data Processing
# ============================================================================

def add_staleness_category(df: pd.DataFrame, last_accessed_col: str) -> pd.DataFrame:
    now = datetime.now(timezone.utc)

    def days_since(x):
        if pd.isna(x):
            return None
        dt = x.astimezone(timezone.utc) if x.tzinfo else x.replace(tzinfo=timezone.utc)
        return max((now - dt).days, 0)

    df["days_since_last_access"] = df[last_accessed_col].apply(days_since)
    df["staleness_category"] = pd.cut(
        df["days_since_last_access"], bins=STALENESS_BINS,
        labels=STALENESS_LABELS, right=True,
    ).astype(str)  # Convert from Categorical to str — allows "Never Accessed" assignment
    df.loc[df["days_since_last_access"].isna(), "staleness_category"] = "Never Accessed"
    # pd.cut assigns "nan" string to NaN values after astype(str)
    df.loc[df["staleness_category"] == "nan", "staleness_category"] = "Never Accessed"
    df["archive_candidate"] = df[last_accessed_col].isna() | (df["days_since_last_access"] > 90)
    df["archive_candidate"] = df["archive_candidate"].map({True: "Yes", False: "No"})
    return df


def add_hierarchy_columns(df: pd.DataFrame, hierarchies: dict,
                          wb_project_map: pd.DataFrame) -> pd.DataFrame:
    """Add top_level_project and project_path via API project ID lookup.

    Merges workbook → api_project_id (from API workbooks endpoint), then
    looks up hierarchy by that ID. This avoids ambiguity when multiple
    projects share the same name at different hierarchy levels.
    """
    # Step 1: Build flat lookup from ID-keyed hierarchies
    rows = []
    for site_url, id_lookup in hierarchies.items():
        for proj_id, data in id_lookup.items():
            rows.append({
                "site_content_url": site_url,
                "api_project_id": proj_id,
                "top_level_project": data["top_level_project"],
                "project_path": data["project_path"],
                "project_depth": data["project_depth"],
            })

    if not rows or wb_project_map.empty:
        df["top_level_project"] = ""
        df["project_path"] = ""
        df["project_depth"] = 0
        return df

    df_hier = pd.DataFrame(rows)

    # Step 2: Attach API project ID to workbooks via (site, workbook_name, project_name)
    df = df.merge(
        wb_project_map[["site_content_url", "workbook_name", "project_name", "api_project_id"]],
        on=["site_content_url", "workbook_name", "project_name"],
        how="left",
    )

    # Step 3: Look up hierarchy by (site, api_project_id)
    df = df.merge(
        df_hier,
        on=["site_content_url", "api_project_id"],
        how="left",
    )

    df.drop(columns=["api_project_id"], inplace=True)
    df[["top_level_project", "project_path"]] = df[["top_level_project", "project_path"]].fillna("")
    df["project_depth"] = df["project_depth"].fillna(0)
    return df


def add_hierarchy_to_projects(df: pd.DataFrame, hierarchies: dict) -> pd.DataFrame:
    """Add hierarchy columns to the Projects DataFrame using name-based lookup.

    For the Projects sheet this is acceptable because each row represents a
    project — the name ambiguity only causes issues when mapping workbooks
    to projects (where the same project name at different levels gives
    different paths). Here, duplicate project names will get the first match,
    which is a minor cosmetic issue on the Projects sheet only.
    """
    rows = []
    for site_url, id_lookup in hierarchies.items():
        seen_names: set[str] = set()
        for proj_data in id_lookup.values():
            name = proj_data["name"]
            if name in seen_names:
                continue  # Skip duplicates — first wins
            seen_names.add(name)
            rows.append({
                "site_content_url": site_url,
                "project_name": name,
                "top_level_project": proj_data["top_level_project"],
                "project_path": proj_data["project_path"],
                "project_depth": proj_data["project_depth"],
            })

    if not rows:
        df["top_level_project"] = ""
        df["project_path"] = ""
        df["project_depth"] = 0
        return df

    df_hier = pd.DataFrame(rows)
    return df.merge(df_hier, on=["site_content_url", "project_name"], how="left").fillna(
        {"top_level_project": "", "project_path": "", "project_depth": 0}
    )


def enrich_workbooks_with_api(df: pd.DataFrame, api_data: dict) -> pd.DataFrame:
    """Add hierarchy and datasource references from API."""
    df = add_hierarchy_columns(df, api_data["project_hierarchy"], api_data["workbook_project_map"])

    df_conn = api_data["workbook_connections"]
    if not df_conn.empty:
        wb_ds = (
            df_conn.groupby(["site_content_url", "project_name", "workbook_name"])
            .agg(
                datasource_names=("datasource_name", lambda x: "; ".join(sorted(set(v for v in x.dropna() if v)))),
                connection_types=("connection_type", lambda x: "; ".join(sorted(set(v for v in x.dropna() if v)))),
            )
            .reset_index()
        )
        df = df.merge(wb_ds, on=["site_content_url", "project_name", "workbook_name"], how="left")
        df[["datasource_names", "connection_types"]] = df[["datasource_names", "connection_types"]].fillna("")
    else:
        df["datasource_names"] = ""
        df["connection_types"] = ""
    return df


def enrich_datasources_with_api(df: pd.DataFrame, api_data: dict) -> pd.DataFrame:
    """Merge repo datasources with API connection details and workbook linkage."""
    df_conn = api_data["workbook_connections"]
    df_types = api_data["datasource_types"]

    # Type info from datasources endpoint
    if not df_types.empty:
        type_cols = ["site_content_url", "datasource_name", "datasource_type", "datasource_has_extracts"]
        df = df.merge(
            df_types[type_cols].drop_duplicates(subset=["site_content_url", "datasource_name"]),
            on=["site_content_url", "datasource_name"], how="left",
        )
        # Convert API boolean to user-friendly Yes/No
        df["datasource_has_extracts"] = df["datasource_has_extracts"].map(
            {True: "Yes", False: "No"}
        ).fillna("")
        # Fill NaN for repo datasources with no matching API type info
        df["datasource_type"] = df["datasource_type"].fillna("")
    else:
        df["datasource_type"] = ""
        df["datasource_has_extracts"] = ""

    # Connection details + workbook linkage in one groupby
    if not df_conn.empty:
        def first_non_empty(s):
            """Return first non-null, non-empty value, or empty string."""
            for val in s:
                if pd.notna(val) and val != "":
                    return val
            return ""

        ds_agg = (
            df_conn.groupby(["site_content_url", "datasource_name"])
            .agg(
                connection_type=("connection_type", lambda x: "; ".join(sorted(set(v for v in x.dropna() if v)))),
                connection_server=("connection_server", first_non_empty),
                connection_database=("connection_database", first_non_empty),
                connection_port=("connection_port", first_non_empty),
                connection_username=("connection_username", first_non_empty),
                workbook_names=("workbook_name", lambda x: "; ".join(sorted(set(v for v in x.dropna() if v)))),
                workbook_count=("workbook_name", "nunique"),
            )
            .reset_index()
        )
        df = df.merge(ds_agg, on=["site_content_url", "datasource_name"], how="left")
        # Fill NaN for repo datasources with no matching API connections
        for col in ["connection_type", "connection_server", "connection_database",
                     "connection_port", "connection_username", "workbook_names"]:
            if col in df.columns:
                df[col] = df[col].fillna("")
        if "workbook_count" in df.columns:
            df["workbook_count"] = df["workbook_count"].fillna(0).astype(int)
    else:
        for col in ["connection_type", "connection_server", "connection_database",
                     "connection_port", "connection_username", "workbook_names"]:
            df[col] = ""
        df["workbook_count"] = 0
    return df


def run_query(query_str: str, conn, sites: tuple) -> pd.DataFrame:
    stmt = text(query_str).bindparams(bindparam("sites", expanding=True))
    result = conn.execute(stmt, {"sites": list(sites)})
    return pd.DataFrame(result.fetchall(), columns=result.keys())


def rename_cols(df: pd.DataFrame, col_map: dict[str, str]) -> pd.DataFrame:
    """Select and rename only columns that exist in the DataFrame."""
    existing = {k: v for k, v in col_map.items() if k in df.columns}
    result = df[list(existing.keys())].rename(columns=existing)  # type: ignore[call-overload]
    # Strip timezone from datetime columns — Excel doesn't support tz-aware datetimes
    # and some pandas/openpyxl version combos crash on write
    for col in result.columns:
        if pd.api.types.is_datetime64_any_dtype(result[col]):
            result[col] = result[col].dt.tz_localize(None) if result[col].dt.tz else result[col]
    return result

# ============================================================================
# Excel Formatting
# ============================================================================

HEADER_FONT = Font(name="Arial", bold=True, color="FFFFFF", size=10)
HEADER_FILL = PatternFill("solid", fgColor="2F5496")
HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
DATA_FONT = Font(name="Arial", size=10)
DATA_ALIGN = Alignment(vertical="top")
PLACEHOLDER_FILL = PatternFill("solid", fgColor="FFF2CC")
THIN_BORDER = Border(
    left=Side(style="thin", color="D9D9D9"), right=Side(style="thin", color="D9D9D9"),
    top=Side(style="thin", color="D9D9D9"), bottom=Side(style="thin", color="D9D9D9"),
)
BA_PLACEHOLDER_COLS = {"Stakeholder", "Status", "Stakeholder Comments", "Representative", "Archiving Process"}


def format_worksheet(ws, df: pd.DataFrame):
    ncols = len(df.columns)
    nrows = len(df) + 1  # +1 for header

    for col_idx in range(1, ncols + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font, cell.fill, cell.alignment, cell.border = HEADER_FONT, HEADER_FILL, HEADER_ALIGN, THIN_BORDER

    for row_idx in range(2, nrows + 1):
        for col_idx in range(1, ncols + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            if df.columns[col_idx - 1] in BA_PLACEHOLDER_COLS:
                cell.fill = PLACEHOLDER_FILL
            cell.font, cell.alignment, cell.border = DATA_FONT, DATA_ALIGN, THIN_BORDER

    for col_idx, col_name in enumerate(df.columns, 1):
        sample_len = max(
            (len(str(v)) if v is not None else 0
             for r in range(1, min(nrows + 1, 52))
             for v in [ws.cell(row=r, column=col_idx).value]),
            default=0,
        )
        ws.column_dimensions[get_column_letter(col_idx)].width = min(sample_len + 3, 40)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions


def create_summary_sheet(wb, dfs: dict, run_time: datetime, api_enabled: bool):
    ws = wb.create_sheet("Summary", 0)
    title_font = Font(name="Arial", bold=True, size=14, color="2F5496")
    section_font = Font(name="Arial", bold=True, size=11, color="2F5496")
    label_font = Font(name="Arial", size=10)
    value_font = Font(name="Arial", bold=True, size=10)
    note_font = Font(name="Arial", size=10, italic=True, color="808080")

    row = 1
    ws.cell(row=row, column=1, value="Tableau Workbook Usage Report").font = title_font
    row += 1
    ws.cell(row=row, column=1, value=f"Generated: {run_time.strftime('%Y-%m-%d %H:%M:%S')}").font = label_font
    row += 1
    ws.cell(row=row, column=1, value=f"Sites: {', '.join(SITES_FILTER)}").font = label_font
    row += 1
    mode = "Repository + REST API" if api_enabled else "Repository only (no API credentials)"
    ws.cell(row=row, column=1, value=f"Mode: {mode}").font = note_font
    row += 2

    df_wb = dfs["workbooks"]
    if not df_wb.empty:
        total_wb = len(df_wb)
        archive_count = int((df_wb["archive_candidate"] == "Yes").sum())
        never_accessed = int(df_wb["days_since_last_access"].isna().sum())

        ws.cell(row=row, column=1, value="Overall Metrics").font = section_font
        row += 1
        for label, value in [
            ("Total Workbooks", total_wb),
            ("Total Views", len(dfs["views"]) if not dfs["views"].empty else 0),
            ("Total Datasources", len(dfs["datasources"]) if not dfs["datasources"].empty else 0),
            ("Total Projects", len(dfs["projects"]) if not dfs["projects"].empty else 0),
            ("", ""),
            ("Archive Candidates (>90d or never accessed)", archive_count),
            ("Archive Percentage", f"{archive_count / total_wb * 100:.1f}%"),
            ("Never Accessed Workbooks", never_accessed),
        ]:
            if label:
                ws.cell(row=row, column=1, value=label).font = label_font
                ws.cell(row=row, column=2, value=value).font = value_font
            row += 1

        # Per-site breakdown
        row += 1
        ws.cell(row=row, column=1, value="Per-Site Breakdown").font = section_font
        row += 1
        for ci, h in enumerate(["Site", "Workbooks", "Archive Candidates", "Archive %", "Never Accessed", "Projects"], 1):
            c = ws.cell(row=row, column=ci, value=h)
            c.font, c.fill, c.alignment = HEADER_FONT, HEADER_FILL, HEADER_ALIGN
        row += 1

        df_proj = dfs["projects"]
        for site in SITES_FILTER:
            sd = df_wb[df_wb["site_content_url"].str.lower() == site.lower()]
            if sd.empty:
                continue
            sa = int((sd["archive_candidate"] == "Yes").sum())
            # Count projects from the projects DataFrame (includes empty projects)
            site_projects = 0
            if not df_proj.empty and "site_content_url" in df_proj.columns:
                site_projects = len(df_proj[df_proj["site_content_url"].str.lower() == site.lower()])
            ws.cell(row=row, column=1, value=site).font = label_font
            ws.cell(row=row, column=2, value=len(sd)).font = value_font
            ws.cell(row=row, column=3, value=sa).font = value_font
            ws.cell(row=row, column=4, value=f"{sa / len(sd) * 100:.1f}%").font = value_font
            ws.cell(row=row, column=5, value=int(sd["days_since_last_access"].isna().sum())).font = value_font
            ws.cell(row=row, column=6, value=site_projects).font = value_font
            row += 1

        # Staleness breakdown
        row += 1
        ws.cell(row=row, column=1, value="Staleness Breakdown").font = section_font
        row += 1
        for ci, h in enumerate(["Category", "Count", "Percentage"], 1):
            c = ws.cell(row=row, column=ci, value=h)
            c.font, c.fill, c.alignment = HEADER_FONT, HEADER_FILL, HEADER_ALIGN
        row += 1
        for cat in STALENESS_LABELS + ["Never Accessed"]:
            count = int((df_wb["staleness_category"] == cat).sum())
            ws.cell(row=row, column=1, value=cat).font = label_font
            ws.cell(row=row, column=2, value=count).font = value_font
            ws.cell(row=row, column=3, value=f"{count / total_wb * 100:.1f}%").font = value_font
            row += 1

    if not api_enabled:
        row += 2
        ws.cell(row=row, column=1, value="API Enrichment Not Available").font = section_font
        row += 1
        for note in [
            "Set TABLEAU_PAT_NAME and TABLEAU_PAT_SECRET to enable:",
            "  - Top Parent Project Name (project hierarchy)",
            "  - Datasource connection details (DSL_Dbclass, DSL_Dbname)",
            "  - Datasource-to-workbook linkage",
        ]:
            ws.cell(row=row, column=1, value=note).font = note_font
            row += 1

    row += 1
    ws.cell(row=row, column=1, value="Yellow-highlighted columns are manual tracking fields for BAs.").font = note_font

    for col, w in {"A": 50, "B": 18, "C": 20, "D": 14, "E": 18, "F": 12}.items():
        ws.column_dimensions[col].width = w

# ============================================================================
# Column Display Names
# ============================================================================

WORKBOOK_COLS = {
    "site_name": "Site Name",
    "site_content_url": "Site Content URL",
    "sitename_workbookname": "SiteName_WorkbookName",
    "workbook_name": "Workbook Name",
    "project_name": "Project Name",
    "top_level_project": "Top Parent Project Name",
    "project_path": "Project Path",
    "owner_username": "Owner Username",
    "owner_display_name": "Owner Display Name",
    "workbook_created_at": "Workbook Created",
    "workbook_updated_at": "Workbook Updated",
    "view_count": "View Count",
    "total_views_all_time": "Total Views (All Time)",
    "workbook_last_accessed": "Last Accessed",
    "days_since_last_access": "Days Since Last Access",
    "staleness_category": "Staleness Category",
    "archive_candidate": "Archive Candidate",
    "datasource_names": "Datasource Names",
    "connection_types": "Connection Types",
    "stakeholder": "Stakeholder",
    "status": "Status",
    "stakeholder_comments": "Stakeholder Comments",
    "representative": "Representative",
    "archiving_process": "Archiving Process",
    "snapshot_date": "Snapshot Date",
}

VIEW_COLS = {
    "site_name": "Site Name",
    "site_content_url": "Site Content URL",
    "workbook_name": "Workbook Name",
    "project_name": "Project Name",
    "owner_username": "Owner Username",
    "owner_display_name": "Owner Display Name",
    "view_name": "View Name",
    "view_title": "View Title",
    "total_views_all_time": "Total Views (All Time)",
    "view_last_accessed": "Last Accessed",
    "days_since_last_access": "Days Since Last Access",
    "staleness_category": "Staleness Category",
    "archive_candidate": "Archive Candidate",
    "snapshot_date": "Snapshot Date",
}

PROJECT_COLS = {
    "site_name": "Site Name",
    "site_content_url": "Site Content URL",
    "project_name": "Project Name",
    "top_level_project": "Top Parent Project Name",
    "project_path": "Project Path",
    "project_depth": "Project Depth",
    "project_owner": "Project Owner",
    "workbook_count": "Workbook Count",
    "project_created_at": "Project Created",
    "snapshot_date": "Snapshot Date",
}

DATASOURCE_COLS = {
    "site_name": "Site Name",
    "site_content_url": "Site Content URL",
    "datasource_name": "Datasource Name",
    "datasource_owner": "Datasource Owner",
    "datasource_owner_display": "Owner Display Name",
    "datasource_type": "DSL_Datasource Type",
    "datasource_has_extracts": "DSL_Datasource Refreshable Extract",
    "connection_type": "DSL_Dbclass",
    "connection_server": "Connection Server",
    "connection_database": "DSL_Dbname",
    "connection_port": "Connection Port",
    "connection_username": "Connection Username",
    "workbook_names": "Used By Workbooks",
    "workbook_count": "Workbook Count",
    "datasource_created_at": "Datasource Created",
    "datasource_updated_at": "Datasource Updated",
    "snapshot_date": "Snapshot Date",
}

# ============================================================================
# Main
# ============================================================================

def main():
    run_timestamp = datetime.now(timezone.utc)
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    logger = setup_logging(OUTPUT_DIR, run_timestamp)

    logger.info("=" * 70)
    logger.info("Tableau Workbook Usage Report")
    logger.info("=" * 70)
    logger.info(f"Run timestamp: {run_timestamp.strftime('%Y-%m-%d %H:%M:%S')}")

    repo_user = os.getenv("TABLEAU_REPO_USER")
    repo_password = os.getenv("TABLEAU_REPO_PASSWORD")
    if not repo_user or not repo_password:
        logger.error("TABLEAU_REPO_USER and TABLEAU_REPO_PASSWORD must both be set.")
        sys.exit(1)

    pat_name = os.getenv("TABLEAU_PAT_NAME")
    pat_secret = os.getenv("TABLEAU_PAT_SECRET")
    api_enabled = bool(pat_name and pat_secret)
    logger.info(f"API enrichment: {'ENABLED' if api_enabled else 'DISABLED (set TABLEAU_PAT_NAME + TABLEAU_PAT_SECRET)'}")

    date_stamp = run_timestamp.strftime("%Y%m%d")
    time_stamp = run_timestamp.strftime("%H%M%S")
    output_file = os.path.join(OUTPUT_DIR, f"tableau_usage_report_{date_stamp}_{time_stamp}.xlsx")

    # Warn about existing same-day reports (non-critical — don't crash on failure)
    try:
        existing = [f for f in os.listdir(OUTPUT_DIR)
                    if f.startswith(f"tableau_usage_report_{date_stamp}") and f.endswith(".xlsx")]
        if existing:
            logger.warning(f"Previous same-day report(s) found: {', '.join(existing)}")
    except OSError:
        pass

    # --- Repository ---
    # Use URL.create() so the password is masked in exception tracebacks and logs
    connection_url = URL.create(
        drivername="postgresql+psycopg2",
        username=repo_user,
        password=repo_password,
        host=TABLEAU_REPO_HOST,
        port=TABLEAU_REPO_PORT,
        database=TABLEAU_REPO_DB,
    )
    logger.info(f"Connecting to repository at {TABLEAU_REPO_HOST}:{TABLEAU_REPO_PORT} as {repo_user}")

    try:
        engine = create_engine(connection_url, connect_args={"connect_timeout": 30})
        with engine.connect() as conn:
            conn.execute(text("SELECT 1"))
        logger.info("Repository connected.")
    except Exception as e:
        logger.error(f"Repository connection failed: {e}")
        sys.exit(1)

    query_failures = []
    dfs = {"workbooks": pd.DataFrame(), "views": pd.DataFrame(),
           "projects": pd.DataFrame(), "datasources": pd.DataFrame()}

    queries = [
        ("workbooks", "workbook usage", WORKBOOK_QUERY),
        ("views", "view usage", VIEW_QUERY),
        ("projects", "projects", PROJECT_QUERY),
        ("datasources", "datasources", DATASOURCE_QUERY),
    ]

    try:
        with engine.connect() as conn:
            for key, label, query in queries:
                try:
                    logger.info(f"Querying {label}...")
                    dfs[key] = run_query(query, conn, SITES_FILTER)
                    logger.info(f"  {len(dfs[key])} rows")
                except Exception as e:
                    logger.error(f"{label.title()} query failed: {e}")
                    query_failures.append(key)
                    conn.rollback()
    finally:
        engine.dispose()
        logger.info("Repository connection closed.")

    # --- REST API ---
    api_data = None
    if api_enabled:
        logger.info("-" * 50)
        logger.info("REST API ENRICHMENT")
        logger.info("-" * 50)
        try:
            # pat_name and pat_secret are guaranteed non-None by api_enabled check
            api_data = fetch_api_data(
                TABLEAU_SERVER_URL, TABLEAU_API_VERSION,
                pat_name,   # type: ignore[arg-type]
                pat_secret, # type: ignore[arg-type]
                SITES_FILTER, logger,
            )
            logger.info("API enrichment complete.")
        except Exception as e:
            logger.error(f"API enrichment failed (continuing with repo data only): {e}")
            api_data = None

    # --- Process ---
    if not dfs["workbooks"].empty:
        dfs["workbooks"] = add_staleness_category(dfs["workbooks"], "workbook_last_accessed")
        if api_data:
            dfs["workbooks"] = enrich_workbooks_with_api(dfs["workbooks"], api_data)
        for col in ["stakeholder", "status", "stakeholder_comments", "representative", "archiving_process"]:
            dfs["workbooks"][col] = ""
        dfs["workbooks"]["snapshot_date"] = run_timestamp.strftime("%Y-%m-%d")

    if not dfs["views"].empty:
        dfs["views"] = add_staleness_category(dfs["views"], "view_last_accessed")
        dfs["views"]["snapshot_date"] = run_timestamp.strftime("%Y-%m-%d")

    if not dfs["projects"].empty:
        if api_data:
            dfs["projects"] = add_hierarchy_to_projects(dfs["projects"], api_data["project_hierarchy"])
        dfs["projects"]["snapshot_date"] = run_timestamp.strftime("%Y-%m-%d")

    if not dfs["datasources"].empty:
        if api_data:
            dfs["datasources"] = enrich_datasources_with_api(dfs["datasources"], api_data)
        dfs["datasources"]["snapshot_date"] = run_timestamp.strftime("%Y-%m-%d")

    # --- Write Excel ---
    logger.info(f"Writing: {output_file}")

    sheet_configs = [
        ("Workbook Usage", "workbooks", WORKBOOK_COLS),
        ("View Usage", "views", VIEW_COLS),
        ("Datasources", "datasources", DATASOURCE_COLS),
        ("Projects", "projects", PROJECT_COLS),
    ]

    try:
        with pd.ExcelWriter(output_file, engine="openpyxl") as writer:
            for sheet_name, key, col_map in sheet_configs:
                if not dfs[key].empty:
                    df_out = rename_cols(dfs[key], col_map)
                    df_out.to_excel(writer, sheet_name=sheet_name, index=False)
                    logger.info(f"  {sheet_name}: {len(df_out)} rows, {len(df_out.columns)} cols")

        wb = load_workbook(output_file)
        for sheet_name, key, col_map in sheet_configs:
            if not dfs[key].empty and sheet_name in wb.sheetnames:
                format_worksheet(wb[sheet_name], rename_cols(dfs[key], col_map))

        create_summary_sheet(wb, dfs, run_timestamp, api_enabled)

        # Remove default "Sheet" tab created by openpyxl when no data sheets were written
        if "Sheet" in wb.sheetnames and len(wb.sheetnames) > 1:
            del wb["Sheet"]

        wb.save(output_file)
        logger.info(f"Excel report saved: {output_file}")

        # Copy to fixed-name file for Power BI data source
        latest_file = os.path.join(OUTPUT_DIR, "tableau_usage_report_latest.xlsx")
        shutil.copy2(output_file, latest_file)
        logger.info(f"Power BI source updated: {latest_file}")
    except Exception as e:
        logger.error(f"Failed to write Excel: {e}")
        raise

    # --- Log summary ---
    if not dfs["workbooks"].empty:
        df_wb = dfs["workbooks"]
        total = len(df_wb)
        archive = int((df_wb["archive_candidate"] == "Yes").sum())
        logger.info(f"Total: {total} workbooks, {archive} archive candidates ({archive / total * 100:.1f}%)")
        for site in SITES_FILTER:
            sd = df_wb[df_wb["site_content_url"].str.lower() == site.lower()]
            if not sd.empty:
                sa = int((sd["archive_candidate"] == "Yes").sum())
                logger.info(f"  {site}: {len(sd)} workbooks, {sa} archive ({sa / len(sd) * 100:.1f}%)")

    if query_failures:
        logger.warning(f"Partial completion — failed: {', '.join(query_failures)}")
        sys.exit(1)
    else:
        logger.info("Complete.")


if __name__ == "__main__":
    main()
